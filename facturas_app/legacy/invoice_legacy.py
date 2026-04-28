import os
import re
import uuid
import logging
import multiprocessing
import queue
import pdfplumber
import openpyxl
import shutil
import time
import zipfile
from typing import Optional
import flask
from flask_cors import CORS
import random
from facturas_app.config import get_settings

PROC_DEBUG = False

_settings = get_settings()
BASE_PATH = str(_settings.base_path)
FACTURAS_ROOT = str(_settings.facturas_root)
FACTURAS_PATH = str(_settings.facturas_path)
FACTURAS_PROCESADAS = str(_settings.facturas_procesadas)
EXCEL_SALIDA = str(_settings.excel_salida)

os.makedirs(FACTURAS_PATH, exist_ok=True)
os.makedirs(FACTURAS_PROCESADAS, exist_ok=True)

tiempos_procesamiento = []
archivos_rechazados_global = []
facturas_con_errores_global = []

logger = logging.getLogger(__name__)


def _env_bool(value, default: bool) -> bool:
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


PAGE_TIMEOUT_SECONDS = float(
    os.getenv("PAGE_TIMEOUT_SECONDS", str(_settings.page_timeout_seconds))
)
PAGE_MAX_WORKERS = int(os.getenv("PAGE_MAX_WORKERS", str(_settings.page_max_workers)))
PAGE_TEMP_DIR = os.getenv("PAGE_TEMP_DIR", str(_settings.page_temp_dir))
PAGE_FALLBACK_ENABLED = _env_bool(
    os.getenv("PAGE_FALLBACK_ENABLED"), _settings.page_fallback_enabled
)
PAGE_FALLBACK_LIBRARY = (
    os.getenv("PAGE_FALLBACK_LIBRARY", _settings.page_fallback_library).strip().lower()
)
PAGE_KEEP_TEMP_FILES = _env_bool(
    os.getenv("PAGE_KEEP_TEMP_FILES"), _settings.page_keep_temp_files
)


def mover_archivo_seguro(origen: str, destino: str, max_intentos: int = 5) -> bool:
    for intento in range(max_intentos):
        try:
            shutil.move(origen, destino)
            print(f"    [OK] Archivo movido exitosamente en intento {intento + 1}")
            return True
        except (PermissionError, OSError) as e:
            print(f"    [WARN] Intento {intento + 1}/{max_intentos} falló: {e}")
            if intento < max_intentos - 1:
                tiempo_espera = random.uniform(1.0, 3.0) + (intento * 0.5)
                print(
                    f"    [WAIT] Esperando {tiempo_espera:.1f} segundos antes del siguiente intento..."
                )
                time.sleep(tiempo_espera)

                try:
                    import gc

                    gc.collect()
                except:
                    pass
            else:
                print(
                    f"    [ERROR] No se pudo mover el archivo después de {max_intentos} intentos"
                )
                print(
                    f"    [INFO] Sugerencia: Cierra cualquier programa que pueda estar usando el archivo"
                )
                return False
    return False


def normalizar_numero(valor: str) -> str:
    if not valor:
        return ""
    s = valor.strip()
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "")
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
        return s
    if "," in s:
        parte_decimal = s.split(",")[-1]
        if len(parte_decimal) in (2, 3):
            return s.replace(",", ".")
        return s.replace(",", "")
    return s


def formatear_numero(valor: str) -> str:
    try:
        n = float(valor)
        valor_str = str(valor).replace(",", ".")
        if "." in valor_str:
            decimales = len(valor_str.split(".")[1])
            if decimales >= 3:
                return f"{n:,.3f}".replace(",", "X").replace(".", ",").replace("X", ".")

        return f"{n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return valor


def _extraer_numero_factura(texto: str) -> Optional[str]:
    patrones = [
        r"FACTURA\s*(?:ELECTR[ÓO]NICA\s*DE\s*VENTA)?\s*No\.?\s*([A-Z0-9\-]+)",
        r"FACTURA\.?\s*N[°o]\s*([A-Z0-9\-]+)",
        r"FACTURA\s*DE\s*VENTA\s*No\.?\s*([A-Z0-9\-]+)",
        r"FACTURA\s+([A-Z0-9\-]+)",
    ]
    for patron in patrones:
        m = re.search(patron, texto, flags=re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return None


def _normalizar_nombre_cliente(linea: str) -> str:
    linea = re.sub(r"\bGREEN\s*POINT\b", "", linea, flags=re.IGNORECASE)
    up = linea.upper()
    up = re.sub(r"[^A-ZÁÉÍÓÚÜÑ\s\.]", "", up)
    up = re.sub(r"\s+", " ", up).strip()
    return up.title()


def _extraer_cliente(texto: str) -> tuple[str, str, str]:
    lines = [ln.strip() for ln in texto.splitlines()]
    nit_cliente = ""
    nombre_cliente = ""
    cod_cliente = ""

    for i, line in enumerate(lines):
        if "CLIENTE" in line.upper():

            m = re.search(r"CLIENTE\s*:?\s*([\d\.\-]+)", line, flags=re.IGNORECASE)
            if m and not nit_cliente:
                nit_cliente = m.group(1).strip()

            if not nombre_cliente:
                for j in range(i + 1, min(i + 6, len(lines))):
                    ln = lines[j].strip()
                    if not ln:
                        continue
                    if re.search(r"^DESPACHADO\s*A", ln, re.IGNORECASE):
                        break
                    if re.search(r"^COD\.?\s*CLIENTE", ln, re.IGNORECASE):
                        break
                    if re.search(r"^FECHA", ln, re.IGNORECASE):
                        break
                    if re.match(
                        r"^(Tel|Tel\.|Teléfono|Telefono|Email|Mail|Direcci[oó]n|www\.)",
                        ln,
                        re.IGNORECASE,
                    ):
                        continue
                    if re.match(r"^\d", ln):
                        continue
                    nombre_cliente = _normalizar_nombre_cliente(ln)
                    break

        if "COD. CLIENTE" in line.upper():
            m = re.search(r"COD\.?\s*CLIENTE\s*([\d]+)", line, flags=re.IGNORECASE)
            if m:
                cod_cliente = m.group(1).strip()

    return nit_cliente, nombre_cliente, cod_cliente


def _extraer_fecha_generacion(texto: str) -> str:
    m = re.search(
        r"FECHA\s*GENERACI[ÓO]N.*?(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})",
        texto,
        flags=re.IGNORECASE,
    )
    if m:
        fecha_completa = m.group(1)
        return fecha_completa.split()[0]

    m = re.search(
        r"FECHA\s*GENERACI[ÓO]N.*?(\d{2}/\d{2}/\d{4})", texto, flags=re.IGNORECASE
    )
    if m:
        return m.group(1)

    m = re.search(
        r"FECHA\s*:?\s*(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})",
        texto,
        flags=re.IGNORECASE,
    )
    if m:
        fecha_completa = m.group(1)
        return fecha_completa.split()[0]

    m = re.search(r"(\d{2}/\d{2}/\d{4})", texto)
    return m.group(1) if m else ""


def _extraer_fecha_expedicion(texto: str) -> str:
    print(f"    [DEBUG] Extrayendo fecha de expedición...")
    print(f"    [DEBUG] Texto a analizar (primeros 500 chars): {texto[:500]}")

    def es_contexto_valido(txt):
        txt = txt.upper()
        return not any(x in txt for x in ["LUGAR", "CEDI", "CIUDAD"])

    m = re.search(
        r"FECHA\s*(DE\s*)?EXPEDICI[ÓO]N\s*:?\s*(\d{2}/\d{2}/\d{4})",
        texto,
        flags=re.IGNORECASE,
    )
    if m:
        print(f"    [DEBUG] Caso 1 encontrado: {m.group(2)}")
        return m.group(2)

    m = re.search(
        r"(?<!LUGAR\sDE\s)EXPEDICI[ÓO]N\s*:?\s*(\d{2}/\d{2}/\d{4})",
        texto,
        flags=re.IGNORECASE,
    )
    if m:
        print(f"    [DEBUG] Caso 2 encontrado: {m.group(1)}")
        return m.group(1)

    m = re.search(
        r"FECHA\s+DE\s+EXPEDICI[ÓO]N\s*:?\s*(\d{2}/\d{2}/\d{4})",
        texto,
        flags=re.IGNORECASE,
    )
    if m:
        print(f"    [DEBUG] Caso 3 encontrado: {m.group(1)}")
        return m.group(1)

    lineas = texto.splitlines()
    for i, linea in enumerate(lineas):
        linea_upper = linea.upper()

        if "EXPEDICI" in linea_upper and "LUGAR" not in linea_upper:
            print(f"    [DEBUG] Línea candidata: '{linea}'")

            m = re.search(r"(\d{2}/\d{2}/\d{4})", linea)
            if m:
                print(f"    [DEBUG] Caso 4.1 (misma línea): {m.group(1)}")
                return m.group(1)

            ventana = " ".join(lineas[i : i + 3])
            print(f"    [DEBUG] Ventana analizada: '{ventana}'")

            m = re.search(r"(\d{2}/\d{2}/\d{4})", ventana)
            if m and es_contexto_valido(ventana):
                print(f"    [DEBUG] Caso 4.2 (ventana): {m.group(1)}")
                return m.group(1)

    m = re.search(
        r"(?<!LUGAR\sDE\s)(EXPEDICI[ÓO]N|EMISI[ÓO]N|EMITIDO|FECHA\s+DE\s+EMISI[ÓO]N)[^\n]{0,50}?(\d{2}/\d{2}/\d{4})",
        texto,
        flags=re.IGNORECASE,
    )
    if m:
        contexto = texto[max(0, m.start() - 50) : m.start()]
        if es_contexto_valido(contexto):
            print(f"    [DEBUG] Caso 5 encontrado: {m.group(2)}")
            return m.group(2)

    print(f"    [DEBUG] No se encontró fecha de expedición")
    return ""


def _limpiar_producto(descripcion: str) -> str:
    descripcion = re.sub(r"^\s*\d+[\.\-\|]*", "", descripcion)
    descripcion = re.sub(r"[^A-Za-z0-9ÁÉÍÓÚÜÑáéíóúüñ\.\-/ ]", " ", descripcion)
    descripcion = re.sub(r"\s+", " ", descripcion).strip()
    return descripcion


def _extract_page_text_pdfplumber(pdf_path: str, page_index: int) -> str:
    with pdfplumber.open(pdf_path) as pdf:
        if page_index < 0 or page_index >= len(pdf.pages):
            raise IndexError("page_index out of range")
        return pdf.pages[page_index].extract_text() or ""


def _extract_page_text_pymupdf(pdf_path: str, page_index: int) -> str:
    try:
        import fitz
    except Exception as exc:
        raise RuntimeError("PyMuPDF not available") from exc

    doc = fitz.open(pdf_path)
    try:
        if page_index < 0 or page_index >= doc.page_count:
            raise IndexError("page_index out of range")
        page = doc.load_page(page_index)
        return page.get_text("text") or ""
    finally:
        doc.close()


def _page_text_worker(
    engine: str, pdf_path: str, page_index: int, result_queue
) -> None:
    started_at = time.perf_counter()
    try:
        if engine == "pdfplumber":
            text = _extract_page_text_pdfplumber(pdf_path, page_index)
        elif engine == "pymupdf":
            text = _extract_page_text_pymupdf(pdf_path, page_index)
        else:
            raise ValueError(f"Unsupported engine: {engine}")
        result_queue.put(
            {
                "status": "OK",
                "text": text,
                "error": "",
                "elapsed": time.perf_counter() - started_at,
                "method": engine,
            }
        )
    except Exception as exc:
        result_queue.put(
            {
                "status": "ERROR",
                "text": "",
                "error": str(exc),
                "elapsed": time.perf_counter() - started_at,
                "method": engine,
            }
        )


def _extract_pages_with_timeout(
    engine: str,
    pdf_path: str,
    page_indices: list[int],
    timeout_seconds: float,
    max_workers: int,
) -> dict[int, dict]:
    ctx = multiprocessing.get_context("spawn")
    results: dict[int, dict] = {}
    pending = list(page_indices)
    active: list[dict] = []

    timeout_seconds = max(0.1, float(timeout_seconds))
    max_workers = max(1, int(max_workers or 1))

    def _start_task(page_index: int) -> dict:
        task_queue = ctx.Queue()
        proc = ctx.Process(
            target=_page_text_worker,
            args=(engine, pdf_path, page_index, task_queue),
        )
        proc.daemon = True
        proc.start()
        return {
            "page_index": page_index,
            "process": proc,
            "queue": task_queue,
            "start": time.perf_counter(),
        }

    while pending or active:
        while pending and len(active) < max_workers:
            active.append(_start_task(pending.pop(0)))

        for task in list(active):
            result = None
            try:
                result = task["queue"].get_nowait()
            except queue.Empty:
                result = None
            except Exception:
                result = None

            if result is not None:
                result["page_index"] = task["page_index"]
                results[task["page_index"]] = result
                task["process"].join(timeout=0)
                task["queue"].close()
                active.remove(task)
                continue

            if not task["process"].is_alive():
                results[task["page_index"]] = {
                    "page_index": task["page_index"],
                    "status": "ERROR",
                    "text": "",
                    "error": "worker_exited_without_result",
                    "elapsed": time.perf_counter() - task["start"],
                    "method": engine,
                }
                task["queue"].close()
                active.remove(task)
                continue

            elapsed = time.perf_counter() - task["start"]
            if elapsed >= timeout_seconds:
                task["process"].terminate()
                task["process"].join(timeout=1)
                results[task["page_index"]] = {
                    "page_index": task["page_index"],
                    "status": "TIMEOUT",
                    "text": "",
                    "error": f"timeout after {timeout_seconds:.2f}s",
                    "elapsed": elapsed,
                    "method": engine,
                }
                task["queue"].close()
                active.remove(task)

        if active:
            time.sleep(0.05)

    return results


def _build_temp_pdf(pdf_path: str, page_indices: list[int], output_path: str) -> bool:
    if not page_indices:
        return False
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    try:
        import fitz
    except Exception as exc:
        logger.error("Fallback PDF builder missing PyMuPDF: %s", exc)
        return False

    src = fitz.open(pdf_path)
    dst = fitz.open()
    try:
        for page_index in page_indices:
            dst.insert_pdf(src, from_page=page_index, to_page=page_index)
        dst.save(output_path)
    finally:
        dst.close()
        src.close()
    return True


def _safe_unlink(path: str) -> None:
    try:
        if os.path.exists(path):
            os.remove(path)
    except Exception:
        return


def _es_factura_valida(pdf_path: str) -> tuple[bool, str]:
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if len(pdf.pages) == 0:
                return False, "El archivo PDF no tiene páginas válidas"

            texto_completo = ""
            paginas_a_revisar = min(3, len(pdf.pages))

            for i in range(paginas_a_revisar):
                texto_pagina = pdf.pages[i].extract_text() or ""
                texto_completo += texto_pagina + "\n"

            texto_completo = texto_completo.upper()

            indicadores_factura = [
                "POSTOBON",
                "GASEOSAS LUX",
                "LUX",
                "FACTURA ELECTRÓNICA DE VENTA",
                "FACTURA DE VENTA",
                "FACTURA",
                "CLIENTE:",
                "COD. CLIENTE",
            ]

            indicadores_encontrados = 0
            indicadores_encontrados_lista = []
            for indicador in indicadores_factura:
                if indicador in texto_completo:
                    indicadores_encontrados += 1
                    indicadores_encontrados_lista.append(indicador)

            print(
                f"    DEBUG: Indicadores encontrados: {indicadores_encontrados}/8 - {indicadores_encontrados_lista}"
            )
            print(f"    DEBUG: Contiene POSTOBON: {'POSTOBON' in texto_completo}")
            print(f"    DEBUG: Contiene LUX: {'LUX' in texto_completo}")

            if indicadores_encontrados < 2:
                return False, "No fue posible procesarse - no parece una factura válida"

            if (
                "FACTURA" not in texto_completo
                and "POSTOBON" not in texto_completo
                and "LUX" not in texto_completo
            ):
                return False, "No fue posible procesarse - no parece una factura válida"

            if not re.search(
                r"FACTURA\s+(?:ELECTR[ÓO]NICA\s+DE\s+VENTA\s+)?No\.?\s*[A-Z0-9\-]+",
                texto_completo,
                re.IGNORECASE,
            ):
                return False, "No fue posible procesarse - no parece una factura válida"

            if not re.search(r"CLIENTE\s*:", texto_completo, re.IGNORECASE):
                return False, "No fue posible procesarse - no parece una factura válida"

            if not re.search(r"\b\d{2,8}\b", texto_completo):
                return False, "No fue posible procesarse - no parece una factura válida"

            if not re.search(
                r"\b(PZA|UNIDAD|SIX|Caja|BOL|UN|U|PZ|CAJA|BOTELLA|BOT)\b",
                texto_completo,
                re.IGNORECASE,
            ):
                return False, "No fue posible procesarse - no parece una factura válida"

            if not re.search(
                r"\b\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?\b", texto_completo
            ):
                return False, "No fue posible procesarse - no parece una factura válida"

            return True, "Archivo válido"

    except Exception as e:
        return False, f"Error al leer el archivo: {str(e)}"


def _extraer_productos(
    texto: str,
) -> list[tuple[str, str, str, str, str, str, str, str, str, str]]:
    productos = []
    lines = texto.splitlines()

    def _to_float_safe(s: str):
        try:
            return float(s)
        except Exception:
            return None

    for s in lines:
        s = s.strip()
        if not s:
            continue

        mref = re.match(r"^(\d{3,6})\b", s)
        m_umv = re.search(r"\b(PZA|UNIDAD|SIX|Caja|BOL)\b", s, flags=re.IGNORECASE)
        if not mref or not m_umv:
            continue

        ref = mref.group(1)
        umv = m_umv.group(1)

        desc_ini = mref.end()
        desc_fin = m_umv.start()
        desc_raw = s[desc_ini:desc_fin].strip()
        descripcion = _limpiar_producto(desc_raw)
        producto = descripcion

        nums = re.findall(r"[\d\.,]+", s[m_umv.end() :])
        nums_norm = [normalizar_numero(n) for n in nums]

        if len(nums_norm) < 3:
            continue

        cantidad = nums_norm[0]
        precio_base_u = nums_norm[1]
        subtotal = nums_norm[2]
        total = nums_norm[-1]

        iva = "0.00"
        subtotal_f = _to_float_safe(subtotal)

        for n in nums_norm[3:]:
            val = _to_float_safe(n)
            if val is not None and 0 <= val <= 100:
                iva = f"{val:.2f}"
                break

        total_f = _to_float_safe(total)
        if subtotal_f is not None and (total_f is None or total_f < subtotal_f - 0.01):
            chosen = None
            max_k = min(3, len(nums_norm))
            for k in range(1, max_k + 1):
                cand = nums_norm[-k]
                cand_f = _to_float_safe(cand)
                if cand_f is None:
                    continue
                if cand_f >= subtotal_f - 0.01:
                    chosen = cand
                    break
            if chosen is None:
                chosen = nums_norm[-1]
            total = chosen

        if descripcion and cantidad and total:
            productos.append(
                (
                    ref,
                    descripcion,
                    umv,
                    cantidad,
                    precio_base_u,
                    iva,
                    total,
                    "OK",
                    producto,
                    "",
                )
            )

    return productos


def extraer_datos_factura(
    pdf_path: str,
    facturas_vistas: set,
    paginas_por_bloque: int = 100,
    modo: str = "acumular",
) -> list[dict]:
    datos = []
    archivo_pdf = os.path.basename(pdf_path)
    inicio_extraccion = time.perf_counter()

    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_paginas = len(pdf.pages)
    except Exception as exc:
        logger.error(
            "Extraction failed to open PDF. file=%s error=%s",
            archivo_pdf,
            exc,
        )
        return datos

    page_indices = list(range(total_paginas))
    timeout_seconds = max(0.1, float(PAGE_TIMEOUT_SECONDS))
    max_workers = max(1, int(PAGE_MAX_WORKERS or 1))

    logger.info(
        "Page extraction start. file=%s pages=%s timeout=%.2fs workers=%s engine=pdfplumber",
        archivo_pdf,
        total_paginas,
        timeout_seconds,
        max_workers,
    )

    primary_results = _extract_pages_with_timeout(
        "pdfplumber",
        pdf_path,
        page_indices,
        timeout_seconds,
        max_workers,
    )

    final_texts: dict[int, str] = {}
    final_status: dict[int, str] = {}
    final_method: dict[int, str] = {}
    final_elapsed: dict[int, float] = {}
    final_error: dict[int, str] = {}

    for page_index in page_indices:
        result = primary_results.get(
            page_index,
            {
                "status": "ERROR",
                "text": "",
                "error": "missing_result",
                "elapsed": 0.0,
                "method": "pdfplumber",
            },
        )
        final_texts[page_index] = result.get("text", "")
        final_status[page_index] = result.get("status", "ERROR")
        final_method[page_index] = result.get("method", "pdfplumber")
        final_elapsed[page_index] = float(result.get("elapsed", 0.0))
        final_error[page_index] = result.get("error", "")

    failed_pages = [
        page_index
        for page_index in page_indices
        if final_status.get(page_index) != "OK"
    ]

    temp_dir = os.path.join(PAGE_TEMP_DIR, os.path.splitext(archivo_pdf)[0])
    created_files: list[str] = []
    problematic_pdf_path = ""

    if failed_pages:
        problematic_pdf_path = os.path.join(temp_dir, "paginas_problematicas.pdf")
        if PAGE_FALLBACK_ENABLED and PAGE_FALLBACK_LIBRARY == "pymupdf":
            if _build_temp_pdf(pdf_path, failed_pages, problematic_pdf_path):
                created_files.append(problematic_pdf_path)
                logger.info(
                    "Page fallback start. file=%s pages=%s engine=pymupdf",
                    archivo_pdf,
                    len(failed_pages),
                )
                fallback_results = _extract_pages_with_timeout(
                    "pymupdf",
                    problematic_pdf_path,
                    list(range(len(failed_pages))),
                    timeout_seconds,
                    max_workers,
                )
                for offset, page_index in enumerate(failed_pages):
                    fallback = fallback_results.get(
                        offset,
                        {
                            "status": "ERROR",
                            "text": "",
                            "error": "missing_result",
                            "elapsed": 0.0,
                            "method": "pymupdf",
                        },
                    )
                    if fallback.get("status") == "OK":
                        final_texts[page_index] = fallback.get("text", "")
                        final_status[page_index] = "RECUPERADA"
                        final_method[page_index] = "pymupdf"
                        final_elapsed[page_index] = float(fallback.get("elapsed", 0.0))
                        final_error[page_index] = ""
                    else:
                        final_status[page_index] = fallback.get("status", "ERROR")
                        final_method[page_index] = "pymupdf"
                        final_elapsed[page_index] = float(fallback.get("elapsed", 0.0))
                        final_error[page_index] = fallback.get("error", "")
            else:
                logger.error(
                    "Page fallback PDF creation failed. file=%s",
                    archivo_pdf,
                )
        else:
            if PAGE_FALLBACK_ENABLED:
                logger.warning(
                    "Page fallback skipped. file=%s library=%s",
                    archivo_pdf,
                    PAGE_FALLBACK_LIBRARY,
                )

    ok_pages = [
        page_index
        for page_index in page_indices
        if final_status.get(page_index) in {"OK", "RECUPERADA"}
    ]
    if ok_pages:
        ok_pdf_path = os.path.join(temp_dir, "procesadas_ok.pdf")
        if _build_temp_pdf(pdf_path, ok_pages, ok_pdf_path):
            created_files.append(ok_pdf_path)

    for page_index in page_indices:
        logger.info(
            "Page extraction result. file=%s page=%s method=%s status=%s elapsed=%.3fs error=%s",
            archivo_pdf,
            page_index + 1,
            final_method.get(page_index, "pdfplumber"),
            final_status.get(page_index, "ERROR"),
            final_elapsed.get(page_index, 0.0),
            final_error.get(page_index, ""),
        )

    if created_files and not PAGE_KEEP_TEMP_FILES:
        for path in created_files:
            _safe_unlink(path)
        try:
            if os.path.isdir(temp_dir) and not os.listdir(temp_dir):
                os.rmdir(temp_dir)
        except Exception:
            pass

    logger.info(
        "Extraction start. file=%s mode=%s pages=%s block_size=%s",
        archivo_pdf,
        modo,
        total_paginas,
        paginas_por_bloque,
    )
    for start in range(0, total_paginas, paginas_por_bloque):
        end = min(start + paginas_por_bloque, total_paginas)
        inicio_bloque = time.perf_counter()
        logger.info(
            "Extraction block start. file=%s block=%s-%s",
            archivo_pdf,
            start + 1,
            end,
        )

        numero_factura = None
        nit_cliente, nombre_cliente, cod_cliente = "", "", ""
        fecha_generacion, fecha_expedicion = "", ""
        productos_totales = []

        def guardar_factura():
            if numero_factura and productos_totales:

                procesar_factura = False
                if modo == "separado":
                    procesar_factura = True
                elif modo == "acumular" and numero_factura not in facturas_vistas:
                    procesar_factura = True
                    facturas_vistas.add(numero_factura)

                if procesar_factura:
                    registros_antes = len(datos)
                    for (
                        ref,
                        prod,
                        umv,
                        unidades,
                        precio_base,
                        iva,
                        total,
                        estado_linea,
                        producto,
                        ml,
                    ) in productos_totales:
                        datos.append(
                            {
                                "id": str(uuid.uuid4())[:8],
                                "numero_factura": numero_factura,
                                "nit_cliente": nit_cliente,
                                "nombre_cliente": nombre_cliente,
                                "cod_cliente": cod_cliente,
                                "fecha_generacion": fecha_generacion,
                                "fecha_expedicion": fecha_expedicion,
                                "referencia": ref,
                                "productos": producto,
                                "umv": umv,
                                "unidades": unidades,
                                "precio_base_unitario": precio_base,
                                "iva": iva,
                                "total": total,
                                "estado": estado_linea,
                            }
                        )

                    logger.info(
                        "Extraction invoice committed. file=%s numero_factura=%s productos=%s registros_agregados=%s",
                        archivo_pdf,
                        numero_factura,
                        len(productos_totales),
                        len(datos) - registros_antes,
                    )
                else:
                    logger.info(
                        "Extraction invoice skipped. file=%s numero_factura=%s reason=duplicate_in_acumular",
                        archivo_pdf,
                        numero_factura,
                    )

        for page_index in range(start, end):
            pagina_actual = page_index + 1
            inicio_pagina = time.perf_counter()
            logger.info(
                "Extraction page start. file=%s page=%s",
                archivo_pdf,
                pagina_actual,
            )

            texto = final_texts.get(page_index, "")

            if not texto.strip():
                logger.info(
                    "Extraction page done. file=%s page=%s status=empty elapsed=%.3fs",
                    archivo_pdf,
                    pagina_actual,
                    time.perf_counter() - inicio_pagina,
                )
                continue

            m = re.search(
                r"FACTURA\s+ELECTR[ÓO]NICA\s+DE\s+VENTA\s+No\.?\s*([A-Z0-9\-]+)",
                texto,
                flags=re.IGNORECASE,
            )
            if m:
                nuevo_numero = m.group(1).strip()
                logger.info(
                    "Extraction field done. file=%s page=%s field=numero_factura value=%s",
                    archivo_pdf,
                    pagina_actual,
                    nuevo_numero,
                )
                if numero_factura and nuevo_numero != numero_factura:
                    logger.info(
                        "Extraction invoice boundary detected. file=%s page=%s previous_numero=%s new_numero=%s",
                        archivo_pdf,
                        pagina_actual,
                        numero_factura,
                        nuevo_numero,
                    )
                    guardar_factura()
                    productos_totales = []
                numero_factura = nuevo_numero

                inicio_cliente = time.perf_counter()
                nit_cliente, nombre_cliente, cod_cliente = _extraer_cliente(texto)
                logger.info(
                    "Extraction field done. file=%s page=%s field=cliente elapsed=%.3fs nit=%s cod=%s nombre=%s",
                    archivo_pdf,
                    pagina_actual,
                    time.perf_counter() - inicio_cliente,
                    nit_cliente,
                    cod_cliente,
                    nombre_cliente,
                )

                inicio_fecha_gen = time.perf_counter()
                fecha_generacion = _extraer_fecha_generacion(texto)
                logger.info(
                    "Extraction field done. file=%s page=%s field=fecha_generacion elapsed=%.3fs value=%s",
                    archivo_pdf,
                    pagina_actual,
                    time.perf_counter() - inicio_fecha_gen,
                    fecha_generacion,
                )

                inicio_fecha_exp = time.perf_counter()
                fecha_expedicion = _extraer_fecha_expedicion(texto)
                logger.info(
                    "Extraction field done. file=%s page=%s field=fecha_expedicion elapsed=%.3fs value=%s",
                    archivo_pdf,
                    pagina_actual,
                    time.perf_counter() - inicio_fecha_exp,
                    fecha_expedicion,
                )
                print(f"    [DEBUG] Fecha expedición extraída: '{fecha_expedicion}'")

            inicio_productos = time.perf_counter()
            productos = _extraer_productos(texto)
            logger.info(
                "Extraction field done. file=%s page=%s field=productos elapsed=%.3fs productos_pagina=%s",
                archivo_pdf,
                pagina_actual,
                time.perf_counter() - inicio_productos,
                len(productos),
            )
            if productos:
                productos_totales.extend(productos)
                logger.info(
                    "Extraction page products aggregated. file=%s page=%s productos_acumulados_factura=%s",
                    archivo_pdf,
                    pagina_actual,
                    len(productos_totales),
                )

            logger.info(
                "Extraction page done. file=%s page=%s elapsed=%.3fs",
                archivo_pdf,
                pagina_actual,
                time.perf_counter() - inicio_pagina,
            )

        guardar_factura()
        logger.info(
            "Extraction block done. file=%s block=%s-%s elapsed=%.3fs",
            archivo_pdf,
            start + 1,
            end,
            time.perf_counter() - inicio_bloque,
        )

    logger.info(
        "Extraction done. file=%s elapsed=%.3fs records=%s",
        archivo_pdf,
        time.perf_counter() - inicio_extraccion,
        len(datos),
    )

    return datos


def cargar_facturas_existentes(excel_path: str) -> set:
    facturas_existentes = set()

    if not os.path.exists(excel_path):
        return facturas_existentes

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 1 and row[1]:
                numero_factura = str(row[1]).strip()
                if numero_factura and numero_factura != "None":
                    facturas_existentes.add(numero_factura)

        print(
            f"    [INFO] Cargadas {len(facturas_existentes)} facturas existentes del Excel"
        )
        return facturas_existentes

    except Exception as e:
        print(f"    [WARN] Error al cargar facturas existentes con openpyxl: {e}")

    try:
        import pandas as pd

        df = pd.read_excel(excel_path, engine="openpyxl")
        if not df.empty and len(df.columns) >= 2:
            for _, row in df.iterrows():
                if pd.notna(row.iloc[1]):
                    numero_factura = str(row.iloc[1]).strip()
                    if numero_factura and numero_factura != "None":
                        facturas_existentes.add(numero_factura)
            print(
                f"    [INFO] Cargadas {len(facturas_existentes)} facturas existentes usando pandas"
            )
            return facturas_existentes
    except Exception as e:
        print(f"    [WARN] Error al cargar facturas existentes con pandas: {e}")

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 1 and row[1]:
                numero_factura = str(row[1]).strip()
                if numero_factura and numero_factura != "None":
                    facturas_existentes.add(numero_factura)

        wb.close()
        print(
            f"    [INFO] Cargadas {len(facturas_existentes)} facturas existentes en modo read_only"
        )
        return facturas_existentes

    except Exception as e:
        print(f"    [WARN] Error al cargar facturas existentes en modo read_only: {e}")

    print(
        f"    [WARN] No se pudieron cargar facturas existentes, se procesarán todas como nuevas"
    )
    return facturas_existentes


def _recuperar_datos_excel(excel_path: str) -> list[dict]:
    """
    Intenta recuperar datos del Excel existente usando múltiples métodos.
    Retorna lista de diccionarios con los datos recuperados.
    """
    datos_recuperados = []

    try:
        import pandas as pd

        df = pd.read_excel(excel_path, engine="openpyxl")
        if not df.empty and len(df.columns) >= 14:
            print(
                f"    [INFO] Recuperando datos con pandas: {len(df)} filas encontradas"
            )
            for _, row in df.iterrows():
                if pd.notna(row.iloc[1]):
                    datos_recuperados.append(
                        {
                            "id": str(row.iloc[0]) if pd.notna(row.iloc[0]) else "",
                            "numero_factura": (
                                str(row.iloc[1]) if pd.notna(row.iloc[1]) else ""
                            ),
                            "nit_cliente": (
                                str(row.iloc[2]) if pd.notna(row.iloc[2]) else ""
                            ),
                            "cod_cliente": (
                                str(row.iloc[3]) if pd.notna(row.iloc[3]) else ""
                            ),
                            "nombre_cliente": (
                                str(row.iloc[4]) if pd.notna(row.iloc[4]) else ""
                            ),
                            "fecha_generacion": (
                                str(row.iloc[5]) if pd.notna(row.iloc[5]) else ""
                            ),
                            "fecha_expedicion": (
                                str(row.iloc[6]) if pd.notna(row.iloc[6]) else ""
                            ),
                            "referencia": (
                                str(row.iloc[7]) if pd.notna(row.iloc[7]) else ""
                            ),
                            "productos": (
                                str(row.iloc[8]) if pd.notna(row.iloc[8]) else ""
                            ),
                            "umv": str(row.iloc[9]) if pd.notna(row.iloc[9]) else "",
                            "unidades": (
                                str(row.iloc[10]) if pd.notna(row.iloc[10]) else ""
                            ),
                            "precio_base_unitario": (
                                str(row.iloc[11]) if pd.notna(row.iloc[11]) else ""
                            ),
                            "iva": (
                                str(row.iloc[12]) if pd.notna(row.iloc[12]) else "0.00"
                            ),
                            "total": (
                                str(row.iloc[13]) if pd.notna(row.iloc[13]) else ""
                            ),
                            "estado": (
                                str(row.iloc[14])
                                if len(row) > 14 and pd.notna(row.iloc[14])
                                else "OK"
                            ),
                        }
                    )
            print(
                f"    [OK] Se recuperaron {len(datos_recuperados)} registros del Excel existente"
            )
            return datos_recuperados
    except Exception as e:
        print(f"    [WARN] Método pandas falló: {e}")

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active

        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value) if cell.value else "")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 1 and row[1]:
                datos_recuperados.append(
                    {
                        "id": str(row[0]) if row[0] else "",
                        "numero_factura": str(row[1]) if row[1] else "",
                        "nit_cliente": str(row[2]) if len(row) > 2 and row[2] else "",
                        "cod_cliente": str(row[3]) if len(row) > 3 and row[3] else "",
                        "nombre_cliente": (
                            str(row[4]) if len(row) > 4 and row[4] else ""
                        ),
                        "fecha_generacion": (
                            str(row[5]) if len(row) > 5 and row[5] else ""
                        ),
                        "fecha_expedicion": (
                            str(row[6]) if len(row) > 6 and row[6] else ""
                        ),
                        "referencia": str(row[7]) if len(row) > 7 and row[7] else "",
                        "productos": str(row[8]) if len(row) > 8 and row[8] else "",
                        "umv": str(row[9]) if len(row) > 9 and row[9] else "",
                        "unidades": str(row[10]) if len(row) > 10 and row[10] else "",
                        "precio_base_unitario": (
                            str(row[11]) if len(row) > 11 and row[11] else ""
                        ),
                        "iva": str(row[12]) if len(row) > 12 and row[12] else "0.00",
                        "total": str(row[13]) if len(row) > 13 and row[13] else "",
                        "estado": str(row[14]) if len(row) > 14 and row[14] else "OK",
                    }
                )
        wb.close()
        if datos_recuperados:
            print(
                f"    [OK] Se recuperaron {len(datos_recuperados)} registros con openpyxl"
            )
            return datos_recuperados
    except Exception as e:
        print(f"    [WARN] Método openpyxl falló: {e}")

    return datos_recuperados


def guardar_en_excel(datos: list[dict], modo: str = "acumular"):
    excel_path = EXCEL_SALIDA
    datos_existentes = []
    inicio_guardado_excel = time.perf_counter()
    logger.info(
        "Excel write start. mode=%s path=%s incoming_records=%s",
        modo,
        excel_path,
        len(datos),
    )

    if modo == "acumular" and os.path.exists(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

            print(f"    [INFO] Modo acumular: Recuperando datos existentes del Excel")
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 1 and row[1]:
                    datos_existentes.append(
                        {
                            "id": str(row[0]) if row[0] else "",
                            "numero_factura": str(row[1]) if row[1] else "",
                            "nit_cliente": (
                                str(row[2]) if len(row) > 2 and row[2] else ""
                            ),
                            "cod_cliente": (
                                str(row[3]) if len(row) > 3 and row[3] else ""
                            ),
                            "nombre_cliente": (
                                str(row[4]) if len(row) > 4 and row[4] else ""
                            ),
                            "fecha_generacion": (
                                str(row[5]) if len(row) > 5 and row[5] else ""
                            ),
                            "fecha_expedicion": (
                                str(row[6]) if len(row) > 6 and row[6] else ""
                            ),
                            "referencia": (
                                str(row[7]) if len(row) > 7 and row[7] else ""
                            ),
                            "productos": str(row[8]) if len(row) > 8 and row[8] else "",
                            "umv": str(row[9]) if len(row) > 9 and row[9] else "",
                            "unidades": (
                                str(row[10]) if len(row) > 10 and row[10] else ""
                            ),
                            "precio_base_unitario": (
                                str(row[11]) if len(row) > 11 and row[11] else ""
                            ),
                            "iva": (
                                str(row[12]) if len(row) > 12 and row[12] else "0.00"
                            ),
                            "total": str(row[13]) if len(row) > 13 and row[13] else "",
                            "estado": (
                                str(row[14]) if len(row) > 14 and row[14] else "OK"
                            ),
                        }
                    )

            print(
                f"    [INFO] Modo acumular: {len(datos_existentes)} registros existentes recuperados"
            )
            print(
                f"    [INFO] Agregando {len(datos)} nuevos registros al Excel existente"
            )

        except (
            EOFError,
            openpyxl.utils.exceptions.InvalidFileException,
            zipfile.BadZipFile,
            Exception,
        ) as e:
            print(f"    [WARN] Error al cargar Excel existente con openpyxl: {e}")
            print(
                f"    [INFO] Intentando recuperar datos antes de recrear el archivo..."
            )

            datos_existentes = _recuperar_datos_excel(excel_path)

            if datos_existentes:
                print(
                    f"    [OK] Se recuperaron {len(datos_existentes)} registros antes de recrear el archivo"
                )

                try:
                    backup_path = excel_path + f".backup_{int(time.time())}"
                    shutil.copy2(excel_path, backup_path)
                    print(f"    [INFO] Backup creado: {os.path.basename(backup_path)}")
                except Exception as backup_error:
                    print(f"    [WARN] No se pudo crear backup: {backup_error}")

                try:
                    os.remove(excel_path)
                    print(f"    [OK] Archivo corrupto eliminado")
                except Exception as remove_error:
                    print(
                        f"    [WARN] No se pudo eliminar archivo corrupto: {remove_error}"
                    )
            else:
                print(f"    [WARN] No se pudieron recuperar datos del archivo corrupto")

                try:
                    backup_path = excel_path + f".backup_{int(time.time())}"
                    shutil.copy2(excel_path, backup_path)
                    print(
                        f"    [INFO] Backup creado antes de eliminar: {os.path.basename(backup_path)}"
                    )
                except Exception:
                    pass
                try:
                    os.remove(excel_path)
                except Exception:
                    pass

            print(f"    [INFO] Creando nuevo archivo Excel...")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(
                [
                    "ID",
                    "Número Factura",
                    "NIT Cliente",
                    "Cod Cliente",
                    "Nombre Cliente",
                    "Fecha Generación",
                    "Fecha Expedición",
                    "Referencia",
                    "Producto",
                    "UMV",
                    "Unidades",
                    "Precio Base U",
                    "IVA",
                    "Total",
                    "Estado",
                ]
            )
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "ID",
                "Número Factura",
                "NIT Cliente",
                "Cod Cliente",
                "Nombre Cliente",
                "Fecha Generación",
                "Fecha Expedición",
                "Referencia",
                "Producto",
                "UMV",
                "Unidades",
                "Precio Base U",
                "IVA",
                "Total",
                "Estado",
            ]
        )
        if modo == "separado":
            print(f"    [INFO] Modo separado: Creando nuevo archivo Excel")
        else:
            print(
                f"    [INFO] Modo acumular: Creando nuevo archivo Excel (no existe archivo previo)"
            )

    todos_datos = {}

    for d in datos_existentes:
        clave = (
            d.get("numero_factura", ""),
            d.get("referencia", ""),
            d.get("productos", ""),
            d.get("total", ""),
        )
        if clave not in todos_datos:
            todos_datos[clave] = d

    for d in datos:
        clave = (
            d.get("numero_factura", ""),
            d.get("referencia", ""),
            d.get("productos", ""),
            d.get("total", ""),
        )
        todos_datos[clave] = d

    for d in todos_datos.values():
        precio_base = d.get("precio_base_unitario", "")
        umv = d.get("umv", "")

        if umv.upper() == "BOL":
            precio_base = str(precio_base).replace(",", ".")
        else:
            precio_base = formatear_numero(precio_base)

        ws.append(
            [
                d.get("id", ""),
                d.get("numero_factura", ""),
                d.get("nit_cliente", ""),
                d.get("cod_cliente", ""),
                d.get("nombre_cliente", ""),
                d.get("fecha_generacion", ""),
                d.get("fecha_expedicion", ""),
                d.get("referencia", ""),
                d.get("productos", ""),
                d.get("umv", ""),
                d.get("unidades", ""),
                precio_base,
                d.get("iva", "0.00"),
                formatear_numero(d.get("total", "")),
                d.get("estado", ""),
            ]
        )

    logger.info(
        "Excel write rows staged. mode=%s staged_rows=%s",
        modo,
        len(todos_datos),
    )
    inicio_save_excel = time.perf_counter()
    logger.info("Excel write save start. path=%s", excel_path)

    wb.save(excel_path)
    logger.info(
        "Excel write save done. path=%s elapsed=%.3fs",
        excel_path,
        time.perf_counter() - inicio_save_excel,
    )
    logger.info(
        "Excel write done. mode=%s total_elapsed=%.3fs total_rows=%s",
        modo,
        time.perf_counter() - inicio_guardado_excel,
        len(todos_datos),
    )
    print(
        f"    [OK] Excel guardado con {len(todos_datos)} registros totales ({len(datos_existentes)} existentes + {len(datos)} nuevos)"
    )
    return excel_path


def procesar_facturas(modo: str = "acumular"):
    global tiempos_procesamiento, archivos_rechazados_global, facturas_con_errores_global

    tiempos_procesamiento = []
    archivos_rechazados_global = []
    facturas_con_errores_global = []

    inicio_total = time.time()
    datos_todas = []
    facturas_encontradas = False
    facturas_vistas = set()
    tiempos_individuales = []
    archivos_rechazados = []
    facturas_con_errores = []
    excel_path = None

    if modo == "acumular":
        excel_path = EXCEL_SALIDA
        facturas_existentes = cargar_facturas_existentes(excel_path)
        facturas_vistas.update(facturas_existentes)
        print(
            f"    [INFO] Modo acumular: {len(facturas_existentes)} facturas ya procesadas encontradas"
        )
    else:
        facturas_vistas = set()
        print(
            f"    [INFO] Modo separado: Procesando TODAS las facturas de esta sesión independientemente"
        )
        print(
            f"    [INFO] Modo separado: facturas_vistas inicializado vacío (no se cargan facturas existentes)"
        )

    for archivo in os.listdir(FACTURAS_PATH):
        if archivo.lower().endswith(".pdf"):
            facturas_encontradas = True
            ruta_pdf = os.path.join(FACTURAS_PATH, archivo)

            print(f"\nValidando: {archivo}")
            es_valida, mensaje_error = _es_factura_valida(ruta_pdf)
            print(
                f"  Resultado validación: es_valida={es_valida}, mensaje={mensaje_error}"
            )

            if not es_valida:
                print(f"  [REJECTED] RECHAZADO: {mensaje_error}")
                archivos_rechazados.append({"archivo": archivo, "razon": mensaje_error})
                carpeta_rechazados = os.path.join(FACTURAS_ROOT, "rechazados")
                os.makedirs(carpeta_rechazados, exist_ok=True)
                destino_rechazado = os.path.join(carpeta_rechazados, archivo)
                if not mover_archivo_seguro(ruta_pdf, destino_rechazado):
                    print(
                        f"    [WARN] No se pudo mover {archivo} a rechazados, pero se procesará normalmente"
                    )
                continue

            print(f"  [VALID] VÁLIDO: {mensaje_error}")

            inicio_archivo = time.time()
            print(f"  Procesando: {archivo}")

            try:
                datos = extraer_datos_factura(ruta_pdf, facturas_vistas, 100, modo)
                fin_archivo = time.time()

                tiempo_archivo = fin_archivo - inicio_archivo
                tiempos_individuales.append(
                    {"archivo": archivo, "tiempo": tiempo_archivo}
                )

                if datos:
                    print(
                        f"  [OK] {archivo} procesado exitosamente en {tiempo_archivo:.2f} seg ({len(datos)} registros)"
                    )
                    datos_todas.extend(datos)

                    if modo == "acumular":
                        destino_procesado = os.path.join(FACTURAS_PROCESADAS, archivo)
                        if not mover_archivo_seguro(ruta_pdf, destino_procesado):
                            print(
                                f"    [WARN] No se pudo mover {archivo} a procesadas, pero los datos se guardaron correctamente"
                            )
                    else:
                        print(
                            f"    [INFO] Modo separado: Archivo {archivo} permanece en carpeta Facturas para permitir reprocesamiento"
                        )
                else:
                    print(f"  [WARN] {archivo} procesado pero no se extrajeron datos")
                    facturas_con_errores.append(
                        {
                            "archivo": archivo,
                            "razon": "Esta factura no se pudo procesar",
                        }
                    )
                    carpeta_errores = os.path.join(FACTURAS_ROOT, "errores")
                    os.makedirs(carpeta_errores, exist_ok=True)
                    destino_errores = os.path.join(carpeta_errores, archivo)
                    if not mover_archivo_seguro(ruta_pdf, destino_errores):
                        print(
                            f"    [WARN] No se pudo mover {archivo} a errores, pero se registró el error"
                        )

            except Exception as e:
                fin_archivo = time.time()
                tiempo_archivo = fin_archivo - inicio_archivo
                error_msg = f"Error durante el procesamiento: {str(e)}"
                print(f"  [ERROR] {archivo} falló: {error_msg}")

                facturas_con_errores.append(
                    {"archivo": archivo, "razon": "Esta factura no se pudo procesar"}
                )

                carpeta_errores = os.path.join(BASE_PATH, "errores")
                os.makedirs(carpeta_errores, exist_ok=True)
                destino_errores = os.path.join(carpeta_errores, archivo)
                if not mover_archivo_seguro(ruta_pdf, destino_errores):
                    print(
                        f"    [WARN] No se pudo mover {archivo} a errores, pero se registró el error"
                    )

    if not facturas_encontradas:
        print("\n No hay facturas nuevas para procesar en la carpeta.")
        fin_total = time.time()
        tiempo_total_procesamiento = fin_total - inicio_total

        return {
            "excel_path": None,
            "tiempos": [],
            "archivos_rechazados": [],
            "facturas_con_errores": [],
            "total_seconds": tiempo_total_procesamiento,
            "facturas_procesadas": 0,
            "facturas_nuevas": 0,
            "facturas_duplicadas": 0,
        }

    if archivos_rechazados:
        print(f"\n [INFO] ARCHIVOS RECHAZADOS ({len(archivos_rechazados)}):")
        for rechazado in archivos_rechazados:
            print(f"  [REJECTED] {rechazado['archivo']}: {rechazado['razon']}")
        print(
            f"  [INFO] Archivos rechazados movidos a: {os.path.join(FACTURAS_ROOT, 'rechazados')}"
        )

    if facturas_con_errores:
        print(
            f"\n [WARN] FACTURAS CON ERRORES DE PROCESAMIENTO ({len(facturas_con_errores)}):"
        )
        for error in facturas_con_errores:
            print(f"  [ERROR] {error['archivo']}: {error['razon']}")
        print(
            f"  [INFO] Facturas con errores movidas a: {os.path.join(FACTURAS_ROOT, 'errores')}"
        )

    if datos_todas:
        unicos = {}
        facturas_nuevas = 0
        facturas_duplicadas = 0

        for d in datos_todas:
            clave = (d["numero_factura"], d["referencia"], d["productos"], d["total"])
            if clave not in unicos:
                unicos[clave] = d
                if modo == "acumular":
                    if d["numero_factura"] in facturas_vistas:
                        facturas_duplicadas += 1
                    else:
                        facturas_nuevas += 1
                else:
                    facturas_nuevas += 1
            else:
                if modo == "acumular":
                    facturas_duplicadas += 1

        excel_path = guardar_en_excel(list(unicos.values()), modo)

        print(f"\n [OK] PROCESAMIENTO COMPLETADO:")
        print(f"  [INFO] Datos guardados en {excel_path}")
        print(f"  [INFO] Facturas válidas movidas a {FACTURAS_PROCESADAS}")
        print(f"  [INFO] Total de registros únicos: {len(unicos)}")

        print(f"  [INFO] Modo: {modo}")

        tiempos_procesamiento = tiempos_individuales.copy()
        archivos_rechazados_global = archivos_rechazados.copy()
        facturas_con_errores_global = facturas_con_errores.copy()

        if tiempos_individuales:
            print(f"\n [INFO] Resumen de tiempos por archivo:")
            for t in tiempos_individuales:
                print(f"  - {t['archivo']}: {t['tiempo']:.2f} seg")

        fin_total = time.time()
        tiempo_total_procesamiento = fin_total - inicio_total

        return {
            "excel_path": excel_path,
            "tiempos": tiempos_individuales,
            "archivos_rechazados": archivos_rechazados,
            "facturas_con_errores": facturas_con_errores,
            "total_seconds": tiempo_total_procesamiento,
            "facturas_procesadas": len(unicos) if "unicos" in locals() else 0,
            "facturas_nuevas": facturas_nuevas,
            "facturas_duplicadas": facturas_duplicadas,
        }
    else:
        print(f"\n [WARN] No se procesaron facturas válidas.")
        if archivos_rechazados:
            print(
                f"  Todos los archivos fueron rechazados por no ser facturas válidas."
            )
        archivos_rechazados_global = archivos_rechazados.copy()
        facturas_con_errores_global = facturas_con_errores.copy()

        fin_total = time.time()
        tiempo_total_procesamiento = fin_total - inicio_total

        return {
            "excel_path": None,
            "tiempos": [],
            "archivos_rechazados": archivos_rechazados,
            "facturas_con_errores": facturas_con_errores,
            "total_seconds": tiempo_total_procesamiento,
            "facturas_procesadas": 0,
        }


if __name__ == "__main__":
    procesar_facturas()
