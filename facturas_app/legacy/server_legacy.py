from __future__ import annotations

from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS

import os
import threading
import time
import importlib.util
import unicodedata
import pandas as pd
import uuid
import subprocess
from datetime import datetime
from facturas_app.config import get_settings

try:
    import win32com.client
    import pythoncom

    SAP_AVAILABLE = True
except ImportError:
    SAP_AVAILABLE = False
    pythoncom = None
    print(
        "[WARNING] win32com.client no disponible. La automatización SAP no funcionará."
    )

try:
    from pywinauto import Application

    PYWINAUTO_AVAILABLE = True
except ImportError:
    PYWINAUTO_AVAILABLE = False
    print(
        "[WARNING] pywinauto no disponible. La automatización de guardado de Excel no funcionará."
    )
    print("[INFO] Instala con: pip install pywinauto")

_settings = get_settings()
BASE_PATH = str(_settings.base_path)

CODIGOS_PRINCIPALES_PATH = str(_settings.web_assets_path)

FACTURAS_ROOT = str(_settings.facturas_root)
FACTURAS_PATH = str(_settings.facturas_path)
FACTURAS_CODIGO_PATH = str(_settings.facturas_codigo_path)
FACTURAS_RECHAZADOS = str(_settings.facturas_rechazados)
FACTURAS_ERRORES = str(_settings.facturas_errores)

os.makedirs(FACTURAS_PATH, exist_ok=True)
os.makedirs(FACTURAS_RECHAZADOS, exist_ok=True)
os.makedirs(FACTURAS_ERRORES, exist_ok=True)

CARPETA_BASE_DIF = str(_settings.carpeta_base_dif)
DESHABILITAR_ACTUALIZACION_EXCEL = bool(_settings.disable_excel_update)

PORTAFOLIOS_DIF = [
    "PORTAFOLIO ARA",
    "PORTAFOLIO CANAL CLIENTES ESPECIALES",
    "PORTAFOLIO CANAL MODERNO",
    "PORTAFOLIO CANAL PREMIUM",
    "PORTAFOLIO CANAL RESIDENCIAL",
    "PORTAFOLIO CENCOSUD COLOMBIA",
    "PORTAFOLIO D1",
    "PORTAFOLIO DOLLAR CITY",
    "PORTAFOLIO FRISBY",
    "PORTAFOLIO GRUPO CBC",
    "PORTAFOLIO GRUPO EXITO",
    "PORTAFOLIO GRUPO MESOFOODS",
    "PORTAFOLIO KFC",
    "PORTAFOLIO LEÑOS Y CARBON",
    "PORTAFOLIO MR BONO",
    "PORTAFOLIO OXXO",
    "PORTAFOLIO PIZZA HUT",
    "PORTAFOLIO SR WOK",
    "PORTAFOLIO SUBWAY",
    "PORTAFOLIO SURTIMAYORISTA",
    "PORTAFOLIO TOSTAO",
]

ANALISTAS_DIF = [
    "manuela escobar",
]

try:
    import xlsxwriter

    ENGINE_XLS = "xlsxwriter"
except Exception:
    ENGINE_XLS = "openpyxl"


def _load_facturas_module():
    """
    Carga dinámica de Facturas/cod_facturas/main.py
    """
    facturas_main_path = os.path.join(FACTURAS_CODIGO_PATH, "main.py")
    if not os.path.exists(facturas_main_path):
        return None

    spec = importlib.util.spec_from_file_location("main_facturas", facturas_main_path)
    if not spec or not spec.loader:
        return None

    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


main_facturas = _load_facturas_module()
procesar_facturas = (
    getattr(main_facturas, "procesar_facturas", None) if main_facturas else None
)

app = Flask(__name__, static_folder=".")
CORS(app)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024

resultados_globales: dict = {"status": "idle"}

progreso_diferencias = {
    "estado": "inactivo",
    "portafolio_actual": "",
    "progreso": 0,
    "total": 0,
    "resultados": {"con_diferencias": [], "sin_diferencias": []},
    "archivo_generado": "",
    "tiempo_inicio": None,
    "tiempo_total": None,
    "mensaje": "",
}

progreso_listas = {
    "estado": "inactivo",
    "etapa": "",
    "progreso": 0,
    "total_canales": 0,
    "canales_procesados": 0,
    "archivo_generado": "",
    "mensaje": "",
}

progreso_portafolios = {
    "estado": "inactivo",
    "portafolio_actual": "",
    "progreso": 0,
    "total": 0,
    "procesados": 0,
    "resultados": {"exitosos": [], "errores": []},
    "tiempo_inicio": None,
    "tiempo_total": None,
    "mensaje": "",
}

progreso_dsd = {
    "estado": "inactivo",
    "etapa": "",
    "progreso": 0,
    "clientes_procesados": 0,
    "total_filas": 0,
    "archivo_generado": "",
    "mensaje": "",
}


def formatear_tiempo(segundos: float) -> str:
    minutos = int(segundos // 60)
    segs = int(segundos % 60)
    if minutos > 0:
        return f"{minutos} min {segs} seg"
    return f"{segs} seg"


@app.route("/")
def index():
    return send_from_directory(CODIGOS_PRINCIPALES_PATH, "index.html")


@app.route("/styles.css")
def styles():
    return send_from_directory(
        CODIGOS_PRINCIPALES_PATH, "styles.css", mimetype="text/css"
    )


OFFICIAL_FILE_NAME = "LISTA DE PRECIOS POSTOBON V2.1 - 2025.10.01.xlsx"
OFFICIAL_SHEET = "Precios Lista Moderno"
OFFICIAL_CHANNEL_COL = "Canal"
OFFICIAL_CODE_COL_CANDIDATES = [
    "CDPDTO",
    "COD_PRODUCTO",
    "Codigo",
    "CÓDIGO",
    "COD",
    "REF",
    "Referencia",
]
PORTFOLIO_CODE_COL_CANDIDATES = [
    "CDPDTO",
    "COD_PRODUCTO",
    "Codigo",
    "CÓDIGO",
    "COD",
    "REF",
    "Referencia",
]
PORTFOLIO_SHEET_CANDIDATES = ["CRUCE", "Hoja1", "Sheet1"]

CANAL_A_PORTAFOLIO = [
    "ARA",
    "Cencosud",
    "D1",
    "Dollar City",
    "Grupo Éxito",
    "Olímpica",
    "Oxxo",
    "Pricesmart",
    "Surtimayorista",
]


def _normalize(s: str) -> str:
    if not isinstance(s, str):
        return ""
    s = s.strip().lower()
    return "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )


def _find_code_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    cols_norm = {_normalize(c): c for c in df.columns}
    for cand in candidates:
        cn = _normalize(cand)
        if cn in cols_norm:
            return cols_norm[cn]
    for c in df.columns:
        cn = _normalize(c)
        if "cod" in cn or "ref" in cn:
            return c
    return None


def _load_official_df() -> tuple[pd.DataFrame, str, str]:
    def _find_official_file() -> str | None:
        try:
            candidatos: list[tuple[float, str]] = []
            roots: list[str] = []
            if os.path.isdir(CARPETA_BASE_DIF):
                roots.append(CARPETA_BASE_DIF)
            if os.path.isdir(BASE_PATH) and BASE_PATH not in roots:
                roots.append(BASE_PATH)

            for base in roots:
                for root, _dirs, files in os.walk(base):
                    for fname in files:
                        if not fname.lower().endswith(
                            (".xlsx", ".xls", ".xlsm", ".xlsb")
                        ):
                            continue
                        fn_norm = _normalize(fname)
                        if _normalize("lista de precios postobon") in fn_norm:
                            full = os.path.join(root, fname)
                            try:
                                mtime = os.path.getmtime(full)
                            except Exception:
                                mtime = 0
                            candidatos.append((mtime, full))

            if candidatos:
                candidatos.sort(reverse=True)
                return candidatos[0][1]
        except Exception:
            return None
        return None

    path = os.path.join(CARPETA_BASE_DIF, OFFICIAL_FILE_NAME)
    if not os.path.exists(path):
        alt = _find_official_file()
        if not alt:
            raise FileNotFoundError(
                f"No se encontró el archivo oficial en {CARPETA_BASE_DIF}. Verifica nombre o subcarpetas."
            )
        path = alt

    sheet_name_used = OFFICIAL_SHEET
    try:
        df = pd.read_excel(path, sheet_name=OFFICIAL_SHEET)
    except Exception:
        xls = pd.ExcelFile(path)
        target = None

        def _norm(s: str) -> str:
            return _normalize(str(s))

        for sh in xls.sheet_names:
            n = _norm(sh)
            if "precios" in n and "moderno" in n:
                target = sh
                break
        if target is None and xls.sheet_names:
            target = xls.sheet_names[0]
        if target is None:
            raise ValueError("No se pudo determinar la hoja de la lista oficial")
        sheet_name_used = target
        df = pd.read_excel(path, sheet_name=target)

    code_col = _find_code_col(df, OFFICIAL_CODE_COL_CANDIDATES)
    if not code_col:
        try:
            raw = pd.read_excel(path, sheet_name=sheet_name_used, header=None)
            header_row = None
            for i in range(min(80, len(raw))):
                row_vals = [str(v).strip() for v in list(raw.iloc[i].values)]
                row_norm = [_normalize(v) for v in row_vals]
                if _normalize(OFFICIAL_CHANNEL_COL) in row_norm and any(
                    _normalize(c) in row_norm for c in OFFICIAL_CODE_COL_CANDIDATES
                ):
                    header_row = i
                    break
            if header_row is not None:
                cols = list(raw.iloc[header_row].values)
                df = raw.iloc[header_row + 1 :].copy()
                df.columns = cols
                df = df.dropna(how="all")
                if OFFICIAL_CHANNEL_COL not in df.columns:
                    for c in list(df.columns):
                        if _normalize(c) == _normalize(
                            OFFICIAL_CHANNEL_COL
                        ) or "canal" in _normalize(c):
                            df = df.rename(columns={c: OFFICIAL_CHANNEL_COL})
                            break
                code_col = _find_code_col(df, OFFICIAL_CODE_COL_CANDIDATES)
        except Exception:
            pass

    if not code_col:
        raise ValueError("No se encontró columna de código en la lista oficial")

    return df, code_col, path


def _find_portfolio_file_for_channel(channel: str) -> str | None:
    ch_norm = _normalize(channel)
    try:
        for fname in os.listdir(CARPETA_BASE_DIF):
            if not fname.lower().endswith((".xlsx", ".xls")):
                continue
            fn_norm = _normalize(fname)
            if fn_norm.startswith("portafolio") and ch_norm in fn_norm:
                return os.path.join(CARPETA_BASE_DIF, fname)
    except Exception:
        return None
    return None


def _load_portfolio_df(path: str) -> tuple[pd.DataFrame, str]:
    last_exc: Exception | None = None
    for sheet in PORTFOLIO_SHEET_CANDIDATES:
        try:
            df = pd.read_excel(path, sheet_name=sheet)
            code_col = _find_code_col(df, PORTFOLIO_CODE_COL_CANDIDATES)
            if code_col:
                return df, code_col
        except Exception as e:
            last_exc = e
            continue
    raise ValueError(
        f"No se pudo cargar hoja/código del portafolio: {os.path.basename(path)} - {last_exc}"
    )


def _compare_sets(
    official_df: pd.DataFrame,
    official_code_col: str,
    portfolio_df: pd.DataFrame,
    portfolio_code_col: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    off_codes = set(
        str(x).strip() for x in official_df[official_code_col].dropna().astype(str)
    )
    port_codes = set(
        str(x).strip() for x in portfolio_df[portfolio_code_col].dropna().astype(str)
    )

    only_off = sorted(off_codes - port_codes)
    only_port = sorted(port_codes - off_codes)

    df_only_off = pd.DataFrame(
        {
            "Codigo": only_off,
            "Comentario": ["En PRECIOS POSTOBON pero no en LISTA PRECIOS"]
            * len(only_off),
        }
    )
    df_only_port = pd.DataFrame(
        {
            "Codigo": only_port,
            "Comentario": ["En LISTA PRECIOS pero no en PRECIOS POSTOBON"]
            * len(only_port),
        }
    )
    return df_only_off, df_only_port


def _run_listas():
    """
    Validador de listas: NO actualiza ningún Excel, solo compara códigos entre
    la lista oficial y los portafolios por canal.
    """
    global progreso_listas
    try:
        progreso_listas["estado"] = "procesando"
        progreso_listas["etapa"] = "Cargando lista oficial"
        progreso_listas["progreso"] = 0
        progreso_listas["total_canales"] = len(CANAL_A_PORTAFOLIO)
        progreso_listas["canales_procesados"] = 0
        progreso_listas["archivo_generado"] = ""
        progreso_listas["mensaje"] = ""

        official_df, official_code_col, _official_path = _load_official_df()

        canal_col = None
        if OFFICIAL_CHANNEL_COL in official_df.columns:
            canal_col = OFFICIAL_CHANNEL_COL
        else:
            for c in official_df.columns:
                cn = _normalize(c)
                if cn == _normalize(OFFICIAL_CHANNEL_COL) or "canal" in cn:
                    canal_col = c
                    break

        if not canal_col:
            raise ValueError("No se encontró la columna de canal en la lista oficial")

        official_df["_CANAL_NORM"] = (
            official_df[canal_col].astype(str).apply(_normalize)
        )

        salida_path = os.path.join(BASE_PATH, "comparacion_listas.xlsx")
        os.makedirs(os.path.dirname(salida_path), exist_ok=True)

        with pd.ExcelWriter(salida_path, engine=ENGINE_XLS) as writer:
            for i, canal in enumerate(CANAL_A_PORTAFOLIO, 1):
                progreso_listas["etapa"] = f"Procesando canal: {canal}"
                progreso_listas["progreso"] = int(
                    (i - 1) / max(1, len(CANAL_A_PORTAFOLIO)) * 100
                )

                canal_norm = _normalize(canal)
                df_channel = official_df[official_df["_CANAL_NORM"] == canal_norm]
                if df_channel.empty:

                    pd.DataFrame(
                        {
                            "Mensaje": [
                                f"No se encontraron filas para el canal {canal} en la lista oficial"
                            ]
                        }
                    ).to_excel(writer, sheet_name=canal[:31], index=False)
                    progreso_listas["canales_procesados"] += 1
                    continue

                pfile = _find_portfolio_file_for_channel(canal)
                if not pfile:
                    pd.DataFrame(
                        {
                            "Mensaje": [
                                f"No se encontró archivo de portafolio para {canal}"
                            ]
                        },
                    ).to_excel(writer, sheet_name=canal[:31], index=False)
                    progreso_listas["canales_procesados"] += 1
                    continue

                try:
                    port_df, port_code_col = _load_portfolio_df(pfile)
                    df_only_off, df_only_port = _compare_sets(
                        df_channel, official_code_col, port_df, port_code_col
                    )

                    df_only_off.to_excel(
                        writer,
                        sheet_name=f"{canal[:28]}_OFF",
                        index=False,
                    )
                    df_only_port.to_excel(
                        writer,
                        sheet_name=f"{canal[:28]}_PORT",
                        index=False,
                    )
                except Exception as e:
                    pd.DataFrame(
                        {"Mensaje": [f"Error procesando canal {canal}: {str(e)}"]},
                    ).to_excel(writer, sheet_name=canal[:31], index=False)

                progreso_listas["canales_procesados"] += 1

        progreso_listas["estado"] = "completado"
        progreso_listas["etapa"] = "Finalizado"
        progreso_listas["progreso"] = 100
        progreso_listas["archivo_generado"] = salida_path

    except Exception as e:
        progreso_listas["estado"] = "error"
        progreso_listas["etapa"] = "Error"
        progreso_listas["mensaje"] = str(e)


def _wait_sap_not_busy(session, timeout=120):
    """Espera a que SAP termine de procesar."""
    t0 = time.time()
    while time.time() - t0 < timeout:
        try:
            if not session.Busy:
                time.sleep(1.0)

                if not session.Busy:
                    return
        except Exception:
            time.sleep(2)
            return
        time.sleep(0.3)
    raise TimeoutError("SAP sigue procesando y se agotó el tiempo de espera.")


def _connect_sap_session():
    import pythoncom

    pythoncom.CoInitialize()

    if not SAP_AVAILABLE:
        raise ImportError("win32com.client no está disponible. Instala pywin32.")

    sap_gui = None
    try:
        sap_gui = win32com.client.GetObject("SAPGUI")
        print("[CONNECT SAP] Conectado usando GetObject")
    except Exception as e1:
        try:

            sap_gui = win32com.client.GetActiveObject("SAPGUI")
            print("[CONNECT SAP] Conectado usando GetActiveObject")
        except Exception as e2:
            raise RuntimeError(
                f"No se pudo conectar a SAP GUI. Asegúrate de que SAP GUI esté abierto.\n"
                f"Errores: GetObject={e1}, GetActiveObject={e2}"
            )

    if sap_gui is None:
        raise RuntimeError("No se pudo obtener el objeto SAPGUI.")

    application = None

    try:
        if hasattr(sap_gui, "GetScriptingEngine"):
            try:
                application = sap_gui.GetScriptingEngine()
                print("[CONNECT SAP] GetScriptingEngine() como método funcionó")
            except Exception as e:
                print(f"[CONNECT SAP] GetScriptingEngine() falló: {e}")

                try:
                    application = sap_gui.GetScriptingEngine
                    print("[CONNECT SAP] GetScriptingEngine como propiedad funcionó")
                except Exception as e2:
                    print(
                        f"[CONNECT SAP] GetScriptingEngine como propiedad falló: {e2}"
                    )
        else:

            try:
                if hasattr(sap_gui, "Application"):
                    application = sap_gui.Application
                    print("[CONNECT SAP] Application como propiedad funcionó")
            except Exception as e3:
                print(f"[CONNECT SAP] Application falló: {e3}")
    except Exception as e:
        print(f"[CONNECT SAP] Error general obteniendo scripting engine: {e}")

    if application is None:
        try:

            application = win32com.client.Dispatch("SAP.Application")
            print("[CONNECT SAP] Dispatch('SAP.Application') funcionó")
        except Exception as e:
            print(f"[CONNECT SAP] Dispatch falló: {e}")
            raise RuntimeError(
                f"No se pudo obtener el ScriptingEngine de SAP GUI.\n"
                f"Verifica que SAP GUI Scripting esté habilitado en las opciones de SAP.\n"
                f"Error: {e}"
            )

    if application is None:
        raise RuntimeError(
            "No se pudo obtener GetScriptingEngine (application es None)."
        )

    con_count = 0
    try:
        if hasattr(application, "Children"):
            children = application.Children
            if hasattr(children, "Count"):
                con_count = children.Count
            else:

                try:
                    test_conn = children(0)
                    con_count = 1
                except Exception:
                    con_count = 0
    except Exception as e:
        print(f"[CONNECT SAP] Error obteniendo con_count: {e}")
        con_count = 0

    if con_count == 0:

        try:
            connection = application.Children(0)
            session = connection.Children(0)
            if session is None:
                raise RuntimeError(
                    "Se obtuvo session=None desde connection.Children(0)."
                )
            print("[CONNECT SAP] Sesión encontrada (acceso directo)")
            return session
        except Exception as e:
            raise RuntimeError(
                "SAP GUI está abierto pero no se encontró ninguna conexión/sesión.\n"
                "Asegúrate de estar logueada y con una sesión activa.\n"
                f"Detalle: {e}"
            )

    for i in range(con_count):
        try:
            connection = application.Children(i)
            try:
                if hasattr(connection, "Children"):
                    children = connection.Children
                    if hasattr(children, "Count"):
                        ses_count = children.Count
                    else:

                        try:
                            session = children(0)
                            if session is not None:
                                print(
                                    f"[CONNECT SAP] Sesión encontrada en conexión {i}"
                                )
                                return session
                        except Exception:
                            ses_count = 0
                    if ses_count > 0:
                        session = connection.Children(0)
                        if session is None:
                            continue
                        print(
                            f"[CONNECT SAP] Sesión encontrada en conexión {i}, sesión 0"
                        )
                        return session
            except Exception as e:
                print(f"[CONNECT SAP] Error en conexión {i}: {e}")
                continue
        except Exception as e:
            print(f"[CONNECT SAP] Error accediendo a conexión {i}: {e}")
            continue

    raise RuntimeError("No se encontró ninguna sesión activa en SAP (session válida).")


def _go_to_transaction(session, tcode: str):
    """Navega a una transacción SAP."""
    try:
        wnd = session.findById("wnd[0]")
        if hasattr(wnd, "maximize"):
            wnd.maximize()

        okcd = session.findById("wnd[0]/tbar[0]/okcd")
        if hasattr(okcd, "text"):
            okcd.text = f"/n{tcode}"
        else:
            okcd.setText(f"/n{tcode}")

        if hasattr(wnd, "sendVKey"):
            wnd.sendVKey(0)
        else:

            import win32com.client

            shell = win32com.client.Dispatch("WScript.Shell")
            shell.SendKeys("{ENTER}")
    except Exception as e:
        raise RuntimeError(f"No se pudo navegar a la transacción {tcode}. Error: {e}")
    _wait_sap_not_busy(session)


def _safe_get_all_node_keys(tree):
    """Obtiene todas las claves de nodos de forma segura."""
    try:
        if hasattr(tree, "GetAllNodeKeys"):
            return list(tree.GetAllNodeKeys())
        elif hasattr(tree, "getAllNodeKeys"):
            return list(tree.getAllNodeKeys())
        else:
            return []
    except Exception as e:
        print(f"[DEBUG] Error en GetAllNodeKeys: {e}")
        return []


def _safe_get_node_text(tree, key):
    """Obtiene el texto de un nodo de forma segura."""
    try:
        if hasattr(tree, "GetNodeTextByKey"):
            return str(tree.GetNodeTextByKey(key)).strip()
        elif hasattr(tree, "getNodeTextByKey"):
            return str(tree.getNodeTextByKey(key)).strip()
        elif hasattr(tree, "GetItemText"):
            try:
                return str(tree.GetItemText(key, "&Hierarchy")).strip()
            except Exception:
                return str(tree.GetItemText(key)).strip()
        else:
            return ""
    except Exception as e:
        print(f"[DEBUG] Error en GetNodeTextByKey para key {key}: {e}")
        return ""


def _safe_expand_node(tree, key):
    """Expande un nodo de forma segura."""
    try:
        if hasattr(tree, "expandNode"):
            tree.expandNode(key)
            return True
        elif hasattr(tree, "expand"):
            tree.expand(key)
            return True
        else:
            return False
    except Exception as e:
        print(f"[DEBUG] Error expandiendo nodo {key}: {e}")
        return False


def _safe_select_node(tree, key):
    """Selecciona un nodo de forma segura."""
    try:
        if hasattr(tree, "selectNode"):
            tree.selectNode(key)
            return True
        elif hasattr(tree, "select"):
            tree.select(key)
            return True
        else:
            return False
    except Exception as e:
        print(f"[DEBUG] Error seleccionando nodo {key}: {e}")
        return False


def _safe_double_click_node(tree, key):
    """Hace doble clic en un nodo de forma segura."""
    try:
        if hasattr(tree, "doubleClickNode"):
            tree.doubleClickNode(key)
            return True
        elif hasattr(tree, "doubleClick"):
            tree.doubleClick(key)
            return True
        else:
            return False
    except Exception as e:
        print(f"[DEBUG] Error haciendo doble clic en nodo {key}: {e}")
        return False


def _find_tree(session):
    """Obtiene el control del árbol de portafolios."""
    try:
        tree = session.findById(
            "wnd[0]/shellcont/shell/shellcont[0]/shell/shellcont[1]/shell[1]"
        )
        return tree
    except Exception as e:

        alt_paths = [
            "wnd[0]/shellcont/shell/shellcont[0]/shell/shellcont[1]/shell",
            "wnd[0]/shellcont/shell/shellcont[1]/shell",
            "wnd[0]/usr/cntlCONTAINER/shellcont/shell/shellcont[1]/shell",
        ]
        for path in alt_paths:
            try:
                tree = session.findById(path)

                if (
                    hasattr(tree, "GetAllNodeKeys")
                    or hasattr(tree, "selectNode")
                    or hasattr(tree, "getAllNodeKeys")
                ):
                    return tree
            except Exception:
                continue
        raise RuntimeError(
            f"No se pudo encontrar el control del árbol de portafolios. "
            f"Verifica que estés en la transacción correcta. Error original: {e}"
        )


def _find_node_key_by_text(tree, target_text: str) -> str | None:
    """Busca un nodo en el árbol por su texto, expandiendo nodos si es necesario."""
    target = target_text.lower().strip()

    def _expand_all_nodes(tree):
        """Intenta expandir todos los nodos del árbol para hacer visible su contenido."""
        try:
            keys = _safe_get_all_node_keys(tree)
            for k in keys:
                _safe_expand_node(tree, k)
        except Exception:
            pass

    _expand_all_nodes(tree)
    time.sleep(0.5)

    try:
        keys = _safe_get_all_node_keys(tree)
        for k in keys:
            try:
                txt = _safe_get_node_text(tree, k).lower()

                if txt == target or target in txt or txt in target:
                    return k
            except Exception:
                continue
    except Exception:
        pass

    try:
        keys = _safe_get_all_node_keys(tree)
        for k in keys:
            try:
                txt = _safe_get_node_text(tree, k).lower()

                if "lista" in txt or "diferencial" in txt or "precio" in txt:
                    try:
                        _safe_expand_node(tree, k)
                        time.sleep(0.3)

                        keys2 = _safe_get_all_node_keys(tree)
                        for k2 in keys2:
                            try:
                                txt2 = _safe_get_node_text(tree, k2).lower()
                                if txt2 == target or target in txt2 or txt2 in target:
                                    return k2
                            except Exception:
                                continue
                    except Exception:
                        continue
            except Exception:
                continue
    except Exception:
        pass

    return None


def _expand_folder_recursive(tree, key, max_depth=5, current_depth=0):
    """Expande recursivamente un nodo y sus hijos hasta encontrar el objetivo."""
    if current_depth >= max_depth:
        return

    try:

        _safe_expand_node(tree, key)
        time.sleep(0.3)

        try:

            all_keys = _safe_get_all_node_keys(tree)
            for child_key in all_keys:
                try:

                    child_txt = _safe_get_node_text(tree, child_key).lower()

                    if any(
                        word in child_txt
                        for word in [
                            "favorito",
                            "lista",
                            "diferencial",
                            "precio",
                            "reporte",
                        ]
                    ):
                        _expand_folder_recursive(
                            tree, child_key, max_depth, current_depth + 1
                        )
                except Exception:
                    continue
        except Exception:
            pass
    except Exception:
        pass


def _select_portfolio(session, portfolio_text: str):
    """
    Selecciona un portafolio en el árbol lateral de SAP (como Cliente/Material, Gr.client./Material, etc.)
    - No depende de "Lista/diferencial precio" (porque en tu pantalla no aparece así).
    - Busca por coincidencia exacta y luego por coincidencia normalizada/fuzzy.
    - Evita expandir nodos que no son carpeta o que el control no soporta expand.
    """
    import time
    import re

    tree = _find_tree(session)
    if tree is None:
        raise RuntimeError("No se encontró el árbol (tree) en la pantalla de SAP.")

    def _norm(s: str) -> str:
        if s is None:
            return ""
        s = str(s).strip().lower()
        s = s.replace("\\", "/")
        s = re.sub(r"\s+", "", s)
        s = s.replace("..", ".")
        return s

    def _safe_text(tree_obj, k) -> str:
        """
        Intenta obtener el texto del nodo con varias APIs comunes:
        - GetNodeTextByKey
        - GetItemText (con columnas típicas)
        - getNodeText / nodeText / Text
        """

        try:
            if hasattr(tree_obj, "GetNodeTextByKey"):
                try:
                    t = tree_obj.GetNodeTextByKey(k)
                    if t:
                        return str(t)
                except Exception:

                    for col in (
                        "",
                        "&Hierarchy",
                        "Hierarchy",
                        "Column1",
                        "DESCR",
                        "TEXT",
                    ):
                        try:
                            t = tree_obj.GetNodeTextByKey(k, col)
                            if t:
                                return str(t)
                        except Exception:
                            continue
        except Exception:
            pass

        try:
            if hasattr(tree_obj, "GetItemText"):
                for col in ("&Hierarchy", "Hierarchy", "Column1", "TEXT", "DESCR", ""):
                    try:
                        t = tree_obj.GetItemText(k, col)
                        if t:
                            return str(t)
                    except Exception:
                        continue
        except Exception:
            pass
        for attr in ("nodeText", "text", "Text"):
            try:
                if hasattr(tree_obj, attr):
                    t = getattr(tree_obj, attr)
                    if callable(t):
                        t = t()
                    if t:
                        return str(t)
            except Exception:
                continue

        return ""

    def _safe_keys(tree_obj):
        try:
            if hasattr(tree_obj, "GetAllNodeKeys"):
                ks = tree_obj.GetAllNodeKeys()
                return list(ks)
        except Exception:
            pass

        try:
            if hasattr(tree_obj, "NodeCount") and hasattr(
                tree_obj, "GetNodeKeyByIndex"
            ):
                cnt = int(tree_obj.NodeCount)
                out = []
                for i in range(cnt):
                    try:
                        out.append(tree_obj.GetNodeKeyByIndex(i))
                    except Exception:
                        continue
                return out
        except Exception:
            pass

        try:
            return _safe_get_all_node_keys(tree_obj)
        except Exception:
            return []

    def _safe_expand_if_possible(tree_obj, k):
        """
        Solo intenta expandir si:
        - existe expandNode/ExpandNode y
        - el nodo parece carpeta (IsFolder/NodeHasChildren/etc.) o al menos no rompe.
        """
        exp_fn = None
        for name in ("expandNode", "ExpandNode"):
            if hasattr(tree_obj, name):
                exp_fn = getattr(tree_obj, name)
                break
        if exp_fn is None:
            return False

        try:
            if hasattr(tree_obj, "IsFolder"):
                if not tree_obj.IsFolder(k):
                    return False
        except Exception:
            pass

        try:
            exp_fn(k)
            return True
        except Exception:
            return False

    wanted_raw = (portfolio_text or "").strip()
    wanted_norm = _norm(wanted_raw)

    print(f"[SELECT PORTFOLIO] Buscando portafolio: {wanted_raw}")

    keys = _safe_keys(tree)
    if not keys:
        raise RuntimeError("No pude obtener las llaves (keys) del árbol de SAP.")

    nodes = []
    for k in keys:
        try:
            t = _safe_text(tree, k)
            if t and t.strip():
                nodes.append((k, t.strip()))
        except Exception:
            continue

    if len(nodes) < 3:
        for k in keys[:50]:
            _safe_expand_if_possible(tree, k)
        time.sleep(0.3)
        nodes = []
        keys2 = _safe_keys(tree)
        for k in keys2:
            try:
                t = _safe_text(tree, k)
                if t and t.strip():
                    nodes.append((k, t.strip()))
            except Exception:
                continue

    key = None
    wanted_low = wanted_raw.lower()
    for k, t in nodes:
        if t.lower() == wanted_low:
            key = k
            print(f"[SELECT PORTFOLIO] Encontrado (exacto): {t}")
            break

    if key is None:
        for k, t in nodes:
            if _norm(t) == wanted_norm:
                key = k
                print(f"[SELECT PORTFOLIO] Encontrado (normalizado): {t}")
                break

    if key is None:
        parts = [p for p in wanted_norm.split("/") if p]
        for k, t in nodes:
            tn = _norm(t)
            if wanted_norm in tn or tn in wanted_norm:
                key = k
                print(f"[SELECT PORTFOLIO] Encontrado (parcial): {t}")
                break
            if parts and all(p in tn for p in parts[-2:]):
                key = k
                print(f"[SELECT PORTFOLIO] Encontrado (por partes): {t}")
                break

    if key is None:
        muestra = ", ".join([t for _, t in nodes[:15]])
        raise RuntimeError(
            f"No se encontró el portafolio '{wanted_raw}' en el árbol de SAP.\n"
            f"Ejemplos de nodos detectados: {muestra}"
        )

    print(f"[SELECT PORTFOLIO] Seleccionando portafolio...")

    try:
        if hasattr(tree, "ensureVisibleHorizontalItem"):
            try:
                tree.ensureVisibleHorizontalItem(key, "&Hierarchy")
            except Exception:
                try:
                    tree.ensureVisibleHorizontalItem(key)
                except Exception:
                    pass
    except Exception:
        pass

    ok_select = False
    for fn_name in ("selectNode", "SelectNode", "selectItem", "SelectItem"):
        try:
            if hasattr(tree, fn_name):
                getattr(tree, fn_name)(key)
                ok_select = True
                break
        except Exception:
            continue

    if not ok_select:
        try:
            ok_select = _safe_select_node(tree, key)
        except Exception:
            ok_select = False

    if not ok_select:
        raise RuntimeError(
            f"No se pudo seleccionar el portafolio '{wanted_raw}' (key={key})."
        )

    time.sleep(0.25)

    ok_open = False
    for fn_name in (
        "doubleClickNode",
        "DoubleClickNode",
        "doubleClickItem",
        "DoubleClickItem",
    ):
        try:
            if hasattr(tree, fn_name):
                getattr(tree, fn_name)(key)
                ok_open = True
                break
        except Exception:
            continue

    if not ok_open:
        try:
            ok_open = _safe_double_click_node(tree, key)
        except Exception:
            ok_open = False

    if not ok_open:
        try:
            session.findById("wnd[0]").sendVKey(0)
            ok_open = True
        except Exception:
            ok_open = False

    if ok_open:
        print("[SELECT PORTFOLIO] Portafolio seleccionado")
    else:
        raise RuntimeError(
            f"Se seleccionó el nodo pero no se pudo abrir/activar el portafolio '{wanted_raw}'."
        )

    _wait_sap_not_busy(session)


def _debug_list_controls(session, max_depth=2, current_path="wnd[0]", current_depth=0):
    """Función de debug para listar controles disponibles (solo para debugging)."""
    if current_depth >= max_depth:
        return []

    controls = []
    try:
        parent = session.findById(current_path)
        if hasattr(parent, "Children"):
            children = parent.Children
            if hasattr(children, "Count"):
                count = children.Count
                for i in range(min(count, 20)):
                    try:
                        child = children(i)
                        child_path = f"{current_path}/{child.type}"
                        controls.append(child_path)

                        if current_depth < max_depth - 1:
                            controls.extend(
                                _debug_list_controls(
                                    session, max_depth, child_path, current_depth + 1
                                )
                            )
                    except Exception:
                        continue
    except Exception:
        pass

    return controls


def _press_consultar(session):
    """Presiona CONSULTAR de forma robusta, sin depender de un ID fijo."""
    print("[PRESS CONSULTAR] Buscando y presionando CONSULTAR (robusto)...")

    time.sleep(0.6)
    _wait_sap_not_busy(session)
    time.sleep(0.4)

    wnd0 = session.findById("wnd[0]")

    def _iter_children(obj, max_depth=8, depth=0):
        """Itera recursivamente los hijos de un control SAP GUI."""
        if depth >= max_depth:
            return
        try:
            ch = obj.Children
            count = ch.Count
        except Exception:
            return

        for i in range(count):
            try:
                c = ch(i)
                yield c
                yield from _iter_children(c, max_depth=max_depth, depth=depth + 1)
            except Exception:
                continue

    toolbar_candidates = []
    for ctrl in _iter_children(wnd0):
        try:
            if hasattr(ctrl, "pressToolbarButton"):
                toolbar_candidates.append(ctrl)
        except Exception:
            continue

    btn_variants = ["CONSULTAR", "EJECUTAR", "ONLI"]
    for tb in toolbar_candidates:
        for b in btn_variants:
            try:
                tb.pressToolbarButton(b)
                print(
                    f"[PRESS CONSULTAR] pressToolbarButton('{b}') en {getattr(tb,'Id','(sin id)')}"
                )
                time.sleep(1.5)
                _wait_sap_not_busy(session, timeout=180)
                time.sleep(2.0)
                print("  Consulta completada")
                return
            except Exception:
                pass

    for ctrl in _iter_children(wnd0):
        try:
            txt = ""
            tip = ""
            if hasattr(ctrl, "Text"):
                txt = str(ctrl.Text or "")
            if hasattr(ctrl, "Tooltip"):
                tip = str(ctrl.Tooltip or "")
            s = (txt + " " + tip).strip().upper()

            if "CONSULTAR" in s or "EJECUTAR" in s:

                if hasattr(ctrl, "press"):
                    ctrl.press()
                    print(f"[PRESS CONSULTAR] Botón presionado por texto/tooltip: {s}")
                    time.sleep(1.5)
                    _wait_sap_not_busy(session, timeout=180)
                    time.sleep(2.0)
                    print("  Consulta completada")
                    return
                if hasattr(ctrl, "click"):
                    ctrl.click()
                    print(f"[PRESS CONSULTAR] Botón clickeado por texto/tooltip: {s}")
                    time.sleep(1.5)
                    _wait_sap_not_busy(session, timeout=180)
                    time.sleep(2.0)
                    print("  Consulta completada")
                    return
        except Exception:
            continue

    grid_paths_try = [
        "wnd[0]/shellcont/shell/shellcont[1]/shell/shellcont[1]/shell",
        "wnd[0]/shellcont/shell/shellcont[1]/shell/shellcont[0]/shell",
        "wnd[0]/usr",
    ]
    for p in grid_paths_try:
        try:
            c = session.findById(p)
            if hasattr(c, "setFocus"):
                c.setFocus()
            break
        except Exception:
            pass

    try:
        wnd0.sendVKey(0)
        print("[PRESS CONSULTAR] Enter enviado (fallback)")
        time.sleep(1.5)
        _wait_sap_not_busy(session, timeout=180)
        time.sleep(2.0)
        print("  Consulta completada")
        return
    except Exception:
        pass

    try:
        wnd0.sendVKey(8)
        print("[PRESS CONSULTAR] F8 enviado (fallback)")
        time.sleep(1.5)
        _wait_sap_not_busy(session, timeout=180)
        time.sleep(2.0)
        print("  Consulta completada")
        return
    except Exception as e:
        raise RuntimeError(
            f"No se pudo presionar CONSULTAR por ningún método. Detalle final: {e}"
        )


def _walk_sap_children(obj, max_depth=12, _depth=0):
    """Recorre recursivamente los hijos del objeto SAP GUI Scripting."""
    if obj is None:
        return
    yield obj

    if _depth >= max_depth:
        return

    try:
        children = getattr(obj, "Children", None)
        if children is None:
            return

        count = children.Count
        for i in range(count):
            try:
                child = children.Item(i)
                yield from _walk_sap_children(
                    child, max_depth=max_depth, _depth=_depth + 1
                )
            except Exception:
                continue
    except Exception:
        return


def _find_alv_grid(session):
    """
    Encuentra el control del ALV Grid recorriendo wnd[0] completo.
    Devuelve el objeto grid (GuiGridView / GuiShell) que soporte export.
    """
    wnd0 = session.findById("wnd[0]")

    candidates = []
    for obj in _walk_sap_children(wnd0, max_depth=14):
        try:
            if hasattr(obj, "pressToolbarContextButton") and hasattr(
                obj, "selectContextMenuItem"
            ):
                obj_id = getattr(obj, "Id", "")
                obj_type = getattr(obj, "Type", "")
                candidates.append((obj_id, obj_type, obj))
        except Exception:
            continue

    if candidates:
        candidates.sort(
            key=lambda x: (("shell" not in (x[0] or "").lower()), len(x[0] or ""))
        )
        best_id, best_type, best_obj = candidates[0]
        print(f"[EXPORT EXCEL] Grid encontrado: Type={best_type} Id={best_id}")
        return best_obj

    return None


def _export_to_excel(session):
    """
    Exporta a Excel usando el menú estándar de SAP:
    Lista → Exportar → Hoja de cálculo
    Funciona incluso cuando NO hay ALV Grid clásico.
    """
    print("[EXPORT EXCEL] Exportando vía menú SAP...")

    _wait_sap_not_busy(session)
    time.sleep(1)

    try:
        session.findById("wnd[0]/mbar/menu[0]").select()
        time.sleep(0.3)

        session.findById("wnd[0]/mbar/menu[0]/menu[3]").select()
        time.sleep(0.3)

        session.findById("wnd[0]/mbar/menu[0]/menu[3]/menu[1]").select()

        print("[EXPORT EXCEL] Exportación a Excel solicitada (menú SAP)")

    except Exception as e:
        raise RuntimeError(f"No se pudo exportar por menú SAP: {e}")

    _wait_sap_not_busy(session)
    time.sleep(1)


def _save_popup_directory_filename(session, directory: str, filename: str):
    """
    Guarda el archivo en el popup de SAP:
    - DY_PATH (carpeta)
    - DY_FILENAME (nombre)
    + confirma reemplazo si aparece
    """
    import os

    directory = os.path.abspath(directory)
    os.makedirs(directory, exist_ok=True)

    print("[SAVE POPUP] Esperando ventana de guardado...")

    t0 = time.time()
    popup_wnd = None
    while time.time() - t0 < 30:
        try:
            popup_wnd = session.findById("wnd[1]")
            print("[SAVE POPUP] Ventana de guardado encontrada")
            break
        except Exception:
            time.sleep(0.2)

    if popup_wnd is None:
        raise TimeoutError("No apareció la ventana de guardado (wnd[1])")

    _wait_sap_not_busy(session)
    time.sleep(0.3)

    wrote_path = False
    wrote_name = False

    for path_id in [
        "wnd[1]/usr/ctxtDY_PATH",
        "wnd[1]/usr/ctxtPATH",
        "wnd[1]/usr/ctxtDYDIR",
    ]:
        try:
            session.findById(path_id).text = directory
            wrote_path = True
            break
        except Exception:
            continue

    for file_id in [
        "wnd[1]/usr/ctxtDY_FILENAME",
        "wnd[1]/usr/ctxtFILENAME",
        "wnd[1]/usr/ctxtDY_FILE",
    ]:
        try:
            file_ctrl = session.findById(file_id)
            try:
                file_ctrl.setFocus()
                file_ctrl.caretPosition = 0
            except Exception:
                pass
            file_ctrl.text = filename
            wrote_name = True
            break
        except Exception:
            continue

    if not wrote_path or not wrote_name:
        raise RuntimeError(
            f"No pude escribir folder/filename en el popup. wrote_path={wrote_path}, wrote_name={wrote_name}"
        )

    time.sleep(0.2)

    saved = False
    for btn_id in ["wnd[1]/tbar[0]/btn[11]", "wnd[1]/tbar[0]/btn[0]"]:
        try:
            session.findById(btn_id).press()
            saved = True
            break
        except Exception:
            continue

    if not saved:
        try:
            popup_wnd.sendVKey(0)
            saved = True
        except Exception:
            pass

    if not saved:
        raise RuntimeError("No se pudo presionar Guardar en el popup (wnd[1]).")

    print(f"[SAVE POPUP] Guardando archivo: {filename}")
    time.sleep(0.4)

    try:
        if session.Children.Count > 1:
            try:
                session.findById("wnd[2]/tbar[0]/btn[0]").press()
                print("[SAVE POPUP] Reemplazo confirmado (wnd[2])")
            except Exception:
                try:
                    session.findById("wnd[1]/tbar[0]/btn[0]").press()
                    print("[SAVE POPUP] Reemplazo confirmado (wnd[1])")
                except Exception:
                    pass
    except Exception:
        pass

    _wait_sap_not_busy(session)


def _normalizar_texto(texto: str) -> str:
    if not isinstance(texto, str):
        return ""
    texto = texto.strip().lower()
    texto = "".join(
        c
        for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )
    return texto


def _buscar_archivo_portafolio(nombre: str) -> str | None:
    nombre_normalizado = nombre.replace(" ", "").lower()
    try:
        for archivo in os.listdir(CARPETA_BASE_DIF):
            archivo_normalizado = archivo.replace(" ", "").lower()
            if nombre_normalizado in archivo_normalizado and archivo.lower().endswith(
                (".xlsx", ".xls", ".xlsm")
            ):
                return os.path.join(CARPETA_BASE_DIF, archivo)
    except Exception:
        pass
    return None


def _procesar_portafolio(nombre: str, ruta: str) -> pd.DataFrame:
    """
    Lee hoja CRUCE, filtra por ANALISTA y DIFERENCIA != 0.
    NO actualiza el archivo, solo lo lee tal como está.
    """
    try:
        if not os.path.exists(ruta):
            print(f"[DIFERENCIAS] Archivo no encontrado: {ruta}")
            return pd.DataFrame()

        try:
            mtime = os.path.getmtime(ruta)
            print(
                f"[DIFERENCIAS] Leyendo {os.path.basename(ruta)} "
                f"(modificado: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(mtime))})"
            )
        except Exception:
            pass

        time.sleep(0.2)

        df = pd.read_excel(ruta, sheet_name="CRUCE", dtype={"CDPDTO": str})
        if df.empty:
            print(f"[DIFERENCIAS] Archivo {nombre} está vacío")
            return pd.DataFrame()

        if "DIFERENCIA" not in df.columns or "ANALISTA" not in df.columns:
            print(f"[DIFERENCIAS] Archivo {nombre} sin columnas DIFERENCIA/ANALISTA")
            return pd.DataFrame(columns=df.columns)

        df["DIFERENCIA"] = pd.to_numeric(df["DIFERENCIA"], errors="coerce")
        df["ANALISTA_NORMALIZADO"] = df["ANALISTA"].astype(str).apply(_normalizar_texto)
        analistas_buscar = [_normalizar_texto(a) for a in ANALISTAS_DIF]

        df_filtrado = df[
            (df["DIFERENCIA"].notna())
            & (df["DIFERENCIA"] != 0)
            & (df["ANALISTA_NORMALIZADO"].isin(analistas_buscar))
        ].copy()

        if "ANALISTA_NORMALIZADO" in df_filtrado.columns:
            df_filtrado = df_filtrado.drop("ANALISTA_NORMALIZADO", axis=1)

        print(f"[DIFERENCIAS] Diferencias encontradas en {nombre}: {len(df_filtrado)}")
        return df_filtrado
    except Exception as e:
        print(f"[DIFERENCIAS] Error procesando {nombre}: {str(e)}")
        return pd.DataFrame()


def _run_diferencias():
    """
    Procesa todos los portafolios definidos en PORTAFOLIOS_DIF.
    NO actualiza ningún archivo Excel; solo lee y genera un consolidado.
    """
    global progreso_diferencias
    try:
        progreso_diferencias["estado"] = "procesando"
        progreso_diferencias["progreso"] = 0
        progreso_diferencias["resultados"] = {
            "con_diferencias": [],
            "sin_diferencias": [],
        }
        progreso_diferencias["portafolio_actual"] = "Iniciando procesamiento..."
        progreso_diferencias["tiempo_inicio"] = time.time()
        progreso_diferencias["tiempo_total"] = None
        progreso_diferencias["total"] = len(PORTAFOLIOS_DIF)
        progreso_diferencias["mensaje"] = ""

        portafolios_con_diferencias: dict[str, pd.DataFrame] = {}
        portafolios_sin_diferencias: list[str] = []

        total_port = len(PORTAFOLIOS_DIF)

        print("=" * 60)
        print("[DIFERENCIAS] INICIANDO ANÁLISIS DE PORTAFOLIOS")
        print("=" * 60)

        for i, nombre in enumerate(PORTAFOLIOS_DIF, 1):
            progreso_diferencias["portafolio_actual"] = (
                f"Analizando: {nombre} ({i}/{total_port})"
            )
            progreso_diferencias["progreso"] = int((i / total_port) * 90)

            ruta = _buscar_archivo_portafolio(nombre)
            if not ruta:
                print(f"[DIFERENCIAS] No se encontró archivo para: {nombre}")
                portafolios_sin_diferencias.append(nombre)
                continue

            df_dif = _procesar_portafolio(nombre, ruta)
            if df_dif is not None and not df_dif.empty:
                portafolios_con_diferencias[nombre] = df_dif
                print(
                    f"[DIFERENCIAS] Diferencias encontradas en {nombre} "
                    f"({len(df_dif)} registros)"
                )
            else:
                portafolios_sin_diferencias.append(nombre)
                print(f"[DIFERENCIAS] Sin diferencias en {nombre}")

        print("\n" + "=" * 60)
        print(
            f"[DIFERENCIAS] Portafolios con diferencias: {len(portafolios_con_diferencias)}"
        )
        print(
            f"[DIFERENCIAS] Portafolios sin diferencias: {len(portafolios_sin_diferencias)}"
        )
        print("=" * 60 + "\n")

        progreso_diferencias["portafolio_actual"] = "Generando reporte consolidado..."
        progreso_diferencias["progreso"] = 95

        archivo_salida = None
        if portafolios_con_diferencias:
            try:
                archivo_salida = os.path.join(BASE_PATH, "diferencias_total.xlsx")
                os.makedirs(os.path.dirname(archivo_salida), exist_ok=True)

                with pd.ExcelWriter(archivo_salida, engine=ENGINE_XLS) as writer:
                    todas_diferencias: list[pd.DataFrame] = []

                    for nombre, df in portafolios_con_diferencias.items():
                        df2 = df.copy()
                        df2["Portafolio"] = nombre
                        todas_diferencias.append(df2)

                    if todas_diferencias:
                        df_consolidado = pd.concat(todas_diferencias, ignore_index=True)
                        df_consolidado.to_excel(
                            writer, sheet_name="TODAS_DIFERENCIAS", index=False
                        )

                    for nombre, df in portafolios_con_diferencias.items():
                        df.to_excel(writer, sheet_name=nombre[:31], index=False)

            except Exception as e:
                print(f"[DIFERENCIAS] Error generando archivo de salida: {e}")
                archivo_salida = None

        progreso_diferencias["portafolio_actual"] = "Procesamiento completado"
        progreso_diferencias["progreso"] = 100
        progreso_diferencias["resultados"]["con_diferencias"] = list(
            portafolios_con_diferencias.keys()
        )
        progreso_diferencias["resultados"][
            "sin_diferencias"
        ] = portafolios_sin_diferencias
        progreso_diferencias["estado"] = "completado"
        progreso_diferencias["tiempo_total"] = time.time() - (
            progreso_diferencias["tiempo_inicio"] or time.time()
        )
        progreso_diferencias["archivo_generado"] = archivo_salida or ""

    except Exception as e:
        progreso_diferencias["estado"] = "error"
        progreso_diferencias["portafolio_actual"] = "Error"
        progreso_diferencias["mensaje"] = str(e)


def _actualizar_excel_portafolio(ruta_excel: str) -> tuple[bool, str]:
    """
    Abre un archivo Excel, ejecuta RefreshAll para actualizar todas las conexiones,
    guarda el archivo y lo cierra.
    Retorna (exito, mensaje_error)
    """
    print(f"    [PORTAFOLIOS] === INICIANDO ACTUALIZACIÓN DE EXCEL ===")
    print(f"    [PORTAFOLIOS] Ruta del archivo: {ruta_excel}")
    print(f"    [PORTAFOLIOS] Archivo existe: {os.path.exists(ruta_excel)}")

    if not SAP_AVAILABLE:
        print(f"    [PORTAFOLIOS] ERROR: win32com.client no está disponible")
        return False, "win32com.client no está disponible"

    excel_app = None
    workbook = None
    inicio = time.time()

    try:
        print(f"    [PORTAFOLIOS] Inicializando COM...")
        import pythoncom

        try:
            try:
                pythoncom.CoInitializeEx(pythoncom.COINIT_APARTMENTTHREADED)
                print(
                    f"    [PORTAFOLIOS] COM inicializado con CoInitializeEx (ApartmentThreaded)"
                )
            except:
                pythoncom.CoInitialize()
                print(f"    [PORTAFOLIOS] COM inicializado con CoInitialize")
        except Exception as e:
            print(f"    [PORTAFOLIOS] Error inicializando COM: {e}")

        print(f"    [PORTAFOLIOS] Creando instancia de Excel...")
        print(f"    [PORTAFOLIOS] === ENTRANDO AL TRY ===")
        try:

            print(f"    [PORTAFOLIOS] === DESPUÉS DE VERIFICAR PROCESOS ===")

            try:
                print(f"    [PORTAFOLIOS] Intentando obtener Excel existente...")
                excel_app = win32com.client.GetActiveObject("Excel.Application")
                print(
                    f"    [PORTAFOLIOS] Excel ya estaba abierto, usando instancia existente"
                )
            except Exception as e1:
                print(
                    f"    [PORTAFOLIOS] Excel no está abierto (error: {e1}), creando nueva instancia..."
                )
                print(
                    f"    [PORTAFOLIOS] IMPORTANTE: Esto puede tardar 10-30 segundos la primera vez..."
                )
                print(f"    [PORTAFOLIOS] Por favor espera, Excel se está iniciando...")

                try:
                    print(
                        f"    [PORTAFOLIOS] Llamando a DispatchEx... (esperando respuesta...)"
                    )
                    excel_app = win32com.client.DispatchEx("Excel.Application")
                    print(f"    [PORTAFOLIOS] Excel Application creada con DispatchEx")
                except Exception as e2:
                    print(
                        f"    [PORTAFOLIOS] DispatchEx falló ({e2}), intentando Dispatch normal..."
                    )
                    print(
                        f"    [PORTAFOLIOS] Llamando a Dispatch... (esperando respuesta...)"
                    )
                    excel_app = win32com.client.Dispatch("Excel.Application")
                    print(f"    [PORTAFOLIOS] Excel Application creada con Dispatch")
        except Exception as e:
            print(f"    [PORTAFOLIOS] ERROR FATAL al crear/obtener Excel: {e}")
            print(f"    [PORTAFOLIOS] Tipo de error: {type(e)}")
            import traceback

            print(f"    [PORTAFOLIOS] Traceback completo:")
            traceback.print_exc()
            raise

        if excel_app is None:
            raise Exception(
                "No se pudo obtener o crear instancia de Excel después de todos los intentos"
            )

        print(f"    [PORTAFOLIOS] Excel listo para usar")
        print(
            f"    [PORTAFOLIOS] Configurando Excel (Visible=True, ScreenUpdating=True)..."
        )
        try:
            excel_app.Visible = True
            print(f"    [PORTAFOLIOS] Visible = True")
        except Exception as e:
            print(f"    [PORTAFOLIOS] Error configurando Visible: {e}")

        try:
            excel_app.DisplayAlerts = False
            excel_app.ScreenUpdating = True
            excel_app.EnableEvents = True
            excel_app.AskToUpdateLinks = False
            excel_app.AlertBeforeOverwriting = False
            print(f"    [PORTAFOLIOS] Excel configurado")
        except Exception as e:
            print(f"    [PORTAFOLIOS] Error en configuración: {e}")

        print(
            f"    [PORTAFOLIOS] Abriendo archivo Excel: {os.path.basename(ruta_excel)}"
        )
        print(f"    [PORTAFOLIOS] Ruta completa: {ruta_excel}")
        try:
            workbook = excel_app.Workbooks.Open(
                ruta_excel, UpdateLinks=True, ReadOnly=False
            )
            print(f"    [PORTAFOLIOS] Archivo abierto exitosamente")
        except Exception as e:
            print(f"    [PORTAFOLIOS] ERROR al abrir archivo: {e}")
            import traceback

            print(f"    [PORTAFOLIOS] Traceback: {traceback.format_exc()}")
            raise

        print(f"    [PORTAFOLIOS] Activando ventana de Excel...")
        try:
            excel_app.Visible = True
            excel_app.WindowState = -4137

            if workbook.Windows.Count > 0:
                workbook.Windows(1).Visible = True
                workbook.Windows(1).WindowState = -4137

            try:
                import win32gui
                import win32con

                hwnd = excel_app.Hwnd
                win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
                win32gui.SetForegroundWindow(hwnd)
                win32gui.BringWindowToTop(hwnd)
            except:
                pass

            print(f"    [PORTAFOLIOS] Ventana activada - DEBERÍAS VER EXCEL AHORA")
            print(f"    [PORTAFOLIOS] Si no ves Excel, revisa la barra de tareas")
        except Exception as e:
            print(f"    [PORTAFOLIOS] Advertencia al activar ventana: {e}")

        print(f"    [PORTAFOLIOS] Ejecutando RefreshAll...")
        workbook.RefreshAll()

        print(f"    [PORTAFOLIOS] Esperando que termine la actualización...")

        max_intentos = 300
        intento = 0

        while intento < max_intentos:
            try:
                estado_calculo = excel_app.CalculationState

                try:
                    excel_app.CalculateUntilAsyncQueriesDone()
                except:
                    pass

                if estado_calculo == 0:
                    time.sleep(1)
                    if excel_app.CalculationState == 0:
                        print(f"    [PORTAFOLIOS] Actualización completada")
                        break

                time.sleep(1)
                intento += 1
                if intento % 10 == 0:
                    print(f"    [PORTAFOLIOS] Esperando... ({intento}s)")

            except Exception as e:
                print(f"    [PORTAFOLIOS] Error verificando estado: {e}")
                time.sleep(2)
                break

        print(f"    [PORTAFOLIOS] Guardando archivo...")
        workbook.Save()

        time.sleep(1)

        workbook.Close(SaveChanges=False)
        workbook = None

        excel_app.Quit()
        excel_app = None

        pythoncom.CoUninitialize()

        tiempo_total = time.time() - inicio
        print(
            f"    [PORTAFOLIOS] Archivo actualizado y guardado en {tiempo_total:.1f}s"
        )
        return True, ""

    except Exception as e:
        error_msg = str(e)
        print(f"    [PORTAFOLIOS] Error: {error_msg}")
        try:
            if workbook:
                try:
                    workbook.Close(SaveChanges=False)
                except:
                    pass
            if excel_app:
                try:
                    excel_app.Quit()
                except:
                    pass
        except:
            pass
        try:
            pythoncom.CoUninitialize()
        except:
            pass
        return False, error_msg


def _run_portafolios():
    """
    Procesa todos los portafolios definidos en PORTAFOLIOS_DIF.
    Abre cada archivo Excel, ejecuta RefreshAll, y guarda los cambios.
    """
    global progreso_portafolios

    progreso_portafolios["estado"] = "procesando"
    progreso_portafolios["portafolio_actual"] = "Iniciando..."
    progreso_portafolios["progreso"] = 0
    progreso_portafolios["total"] = len(PORTAFOLIOS_DIF)
    progreso_portafolios["procesados"] = 0
    progreso_portafolios["resultados"] = {"exitosos": [], "errores": []}
    progreso_portafolios["tiempo_inicio"] = time.time()
    progreso_portafolios["mensaje"] = ""

    exitosos = []
    errores = []

    try:
        print("\n" + "=" * 60)
        print("[PORTAFOLIOS] === PROCESO INICIADO ===")
        print("[PORTAFOLIOS] Iniciando actualización de portafolios")
        print(f"[PORTAFOLIOS] Total de portafolios a procesar: {len(PORTAFOLIOS_DIF)}")
        print(f"[PORTAFOLIOS] Carpeta base: {CARPETA_BASE_DIF}")
        print("=" * 60)

        for i, nombre_portafolio in enumerate(PORTAFOLIOS_DIF, 1):
            progreso_portafolios["portafolio_actual"] = (
                f"Procesando: {nombre_portafolio}"
            )
            progreso_portafolios["progreso"] = int((i - 1) / len(PORTAFOLIOS_DIF) * 100)

            print(
                f"\n[PORTAFOLIOS] [{i}/{len(PORTAFOLIOS_DIF)}] Buscando archivo para: {nombre_portafolio}"
            )
            print(f"[PORTAFOLIOS] Carpeta base: {CARPETA_BASE_DIF}")
            print(f"[PORTAFOLIOS] Carpeta existe: {os.path.exists(CARPETA_BASE_DIF)}")

            ruta_archivo = _buscar_archivo_portafolio(nombre_portafolio)

            if not ruta_archivo:
                error_msg = f"Archivo no encontrado para {nombre_portafolio}"
                print(f"[PORTAFOLIOS] {error_msg}")
                print(f"[PORTAFOLIOS] Listando archivos en carpeta...")
                try:
                    archivos = os.listdir(CARPETA_BASE_DIF)
                    print(f"[PORTAFOLIOS] Archivos encontrados: {len(archivos)}")
                    for arch in archivos[:5]:
                        print(f"  - {arch}")
                except Exception as e:
                    print(f"[PORTAFOLIOS] Error listando archivos: {e}")
                errores.append({"portafolio": nombre_portafolio, "error": error_msg})
                progreso_portafolios["resultados"]["errores"] = errores
                continue

            print(f"[PORTAFOLIOS] Archivo encontrado: {os.path.basename(ruta_archivo)}")
            print(f"[PORTAFOLIOS] Ruta completa: {ruta_archivo}")
            print(f"[PORTAFOLIOS] Archivo existe: {os.path.exists(ruta_archivo)}")
            print(
                f"[PORTAFOLIOS] === INICIANDO ACTUALIZACIÓN - EXCEL SE ABRIRÁ AHORA ==="
            )

            exito, error_msg = _actualizar_excel_portafolio(ruta_archivo)

            if exito:
                print(f"[PORTAFOLIOS] {nombre_portafolio} actualizado exitosamente")
                exitosos.append(nombre_portafolio)
            else:
                print(
                    f"[PORTAFOLIOS] Error actualizando {nombre_portafolio}: {error_msg}"
                )
                errores.append({"portafolio": nombre_portafolio, "error": error_msg})

            progreso_portafolios["procesados"] = i
            progreso_portafolios["resultados"]["exitosos"] = exitosos
            progreso_portafolios["resultados"]["errores"] = errores

        print("\n" + "=" * 60)
        print(f"[PORTAFOLIOS] Actualización completada")
        print(f"[PORTAFOLIOS] Exitosos: {len(exitosos)}")
        print(f"[PORTAFOLIOS] Errores: {len(errores)}")
        print("=" * 60 + "\n")

        progreso_portafolios["portafolio_actual"] = "Actualización completada"
        progreso_portafolios["progreso"] = 100
        progreso_portafolios["estado"] = "completado"
        progreso_portafolios["tiempo_total"] = time.time() - (
            progreso_portafolios["tiempo_inicio"] or time.time()
        )

    except Exception as e:
        print(f"[PORTAFOLIOS] Error general: {e}")
        progreso_portafolios["estado"] = "error"
        progreso_portafolios["portafolio_actual"] = "Error"
        progreso_portafolios["mensaje"] = str(e)


@app.route("/dsd")
@app.route("/dsd/")
def dsd_page():
    return send_from_directory(CODIGOS_PRINCIPALES_PATH, "dsd.html")


@app.route("/dsd.css")
def dsd_css():
    return send_from_directory(
        CODIGOS_PRINCIPALES_PATH, "dsd.css", mimetype="text/css", max_age=0
    )


RUTA_BASE_JERARQUIA = str(_settings.ruta_base_jerarquia)
RUTA_SALIDA_DSD = str(_settings.ruta_salida_dsd)
DSD_TEMP_PATH = str(_settings.dsd_temp_path)
os.makedirs(DSD_TEMP_PATH, exist_ok=True)

archivo_dsd_subido = None


def _run_dsd():
    """
    Procesa Base Jerarquia.xlsx y genera Solicitantes_SAP.xlsx.
    Extrae datos filtrando por Cliente J1 y copiando Cliente J3.
    """
    global progreso_dsd, archivo_dsd_subido

    archivo_a_procesar = None

    try:
        progreso_dsd["estado"] = "procesando"
        progreso_dsd["etapa"] = "Validando archivo de origen..."
        progreso_dsd["progreso"] = 0
        progreso_dsd["clientes_procesados"] = 0
        progreso_dsd["total_filas"] = 0
        progreso_dsd["archivo_generado"] = ""
        progreso_dsd["mensaje"] = ""

        if archivo_dsd_subido and os.path.exists(archivo_dsd_subido):
            archivo_a_procesar = archivo_dsd_subido
        elif os.path.exists(RUTA_BASE_JERARQUIA):
            archivo_a_procesar = RUTA_BASE_JERARQUIA
        else:
            raise FileNotFoundError(
                "No se encontró ningún archivo Base Jerarquia.xlsx. "
                "Por favor, sube el archivo desde la interfaz."
            )

        progreso_dsd["etapa"] = (
            f"Abriendo archivo: {os.path.basename(archivo_a_procesar)}..."
        )
        progreso_dsd["progreso"] = 10

        excel_file = pd.ExcelFile(archivo_a_procesar)
        df = None
        hoja_usada = None

        for sheet_name in excel_file.sheet_names:
            try:
                temp_df = pd.read_excel(archivo_a_procesar, sheet_name=sheet_name)

                cols_norm = {_normalize(c): c for c in temp_df.columns}

                tiene_j1 = any(
                    "cliente" in cn and "j1" in cn for cn in cols_norm.keys()
                )
                tiene_nombre = "nombre" in cols_norm.keys()
                tiene_j3 = any(
                    "cliente" in cn and "j3" in cn for cn in cols_norm.keys()
                )

                if tiene_j1 and tiene_nombre and tiene_j3:
                    df = temp_df
                    hoja_usada = sheet_name
                    break
            except Exception:
                continue

        if df is None:

            df = pd.read_excel(archivo_a_procesar, sheet_name=0)
            hoja_usada = (
                excel_file.sheet_names[0] if excel_file.sheet_names else "Sheet1"
            )

        progreso_dsd["etapa"] = (
            f"Archivo abierto (hoja: {hoja_usada}). Buscando columnas..."
        )
        progreso_dsd["progreso"] = 20

        cols_norm = {_normalize(c): c for c in df.columns}

        col_j1 = None
        col_nombre = None
        col_nombre2 = None
        col_j3 = None
        for cn, c_orig in cols_norm.items():
            if "cliente" in cn and "j1" in cn:
                col_j1 = c_orig
                break

        for cn, c_orig in cols_norm.items():
            if cn == "nombre" or (
                cn.startswith("nombre") and "2" not in cn and "nombre2" not in cn
            ):
                col_nombre = c_orig
                break

        for cn, c_orig in cols_norm.items():
            if "nombre" in cn and (
                "2" in cn or cn == "nombre2" or cn.startswith("nombre2")
            ):
                col_nombre2 = c_orig
                break

        for cn, c_orig in cols_norm.items():
            if "cliente" in cn and "j3" in cn:
                col_j3 = c_orig
                break

        if not col_j1:
            raise ValueError("No se encontró la columna 'Cliente J1' en el archivo")
        if not col_nombre:
            raise ValueError("No se encontró la columna 'Nombre' en el archivo")
        if not col_j3:
            raise ValueError("No se encontró la columna 'Cliente J3' en el archivo")

        progreso_dsd["etapa"] = "Columnas encontradas. Obteniendo clientes únicos..."
        progreso_dsd["progreso"] = 30

        clientes_j1 = df[col_j1].dropna().unique()
        clientes_j1 = [c for c in clientes_j1 if str(c).strip() != ""]

        if len(clientes_j1) == 0:
            raise ValueError("No se encontraron valores en la columna 'Cliente J1'")

        progreso_dsd["etapa"] = f"Procesando {len(clientes_j1)} clientes J1..."
        progreso_dsd["progreso"] = 40

        bloques = []

        for idx, cliente_j1 in enumerate(clientes_j1):
            progreso_dsd["etapa"] = (
                f"Procesando cliente J1: {cliente_j1} ({idx + 1}/{len(clientes_j1)})"
            )
            progreso_dsd["progreso"] = 40 + int((idx / len(clientes_j1)) * 50)

            df_filtrado = df[df[col_j1] == cliente_j1].copy()

            if df_filtrado.empty:
                continue

            nombre_solicitante = (
                df_filtrado[col_nombre].iloc[0] if len(df_filtrado) > 0 else ""
            )

            nombre2_solicitante = ""
            if col_nombre2:
                nombre2_solicitante = (
                    df_filtrado[col_nombre2].iloc[0] if len(df_filtrado) > 0 else ""
                )

            clientes_j3 = df_filtrado[col_j3].dropna()
            clientes_j3 = [c for c in clientes_j3 if str(c).strip() != ""]
            clientes_j3_unicos = []
            vistos = set()
            for c in clientes_j3:
                c_str = str(c).strip()
                if c_str and c_str not in vistos:
                    clientes_j3_unicos.append(c_str)
                    vistos.add(c_str)

            if clientes_j3_unicos:
                bloques.append(
                    {
                        "cliente_j1": str(cliente_j1).strip(),
                        "nombre": (
                            str(nombre_solicitante).strip()
                            if nombre_solicitante
                            else ""
                        ),
                        "nombre2": (
                            str(nombre2_solicitante).strip()
                            if nombre2_solicitante
                            else ""
                        ),
                        "clientes_j3": clientes_j3_unicos,
                    }
                )

            progreso_dsd["clientes_procesados"] = idx + 1

        if not bloques:
            raise ValueError(
                "No se generaron datos. Verifica que el archivo tenga información válida."
            )

        progreso_dsd["etapa"] = (
            "Generando archivo de salida en formato de bloques de columnas..."
        )
        progreso_dsd["progreso"] = 90

        os.makedirs(RUTA_SALIDA_DSD, exist_ok=True)

        archivo_salida = os.path.join(RUTA_SALIDA_DSD, "Solicitantes_SAP.xlsx")

        archivo_generado_exitosamente = False

        try:
            from openpyxl import Workbook

            print(f"[DSD] Intentando generar archivo con openpyxl...")

            wb = Workbook()
            ws = wb.active
            ws.title = "Solicitantes"

            columna_actual = 1

            for idx_bloque, bloque in enumerate(bloques):

                ws.cell(row=1, column=columna_actual, value="Cliente_J1")
                ws.cell(row=1, column=columna_actual + 1, value="Nombre_Solicitante")
                ws.cell(row=1, column=columna_actual + 2, value="Nombre_2")
                ws.cell(row=1, column=columna_actual + 3, value="Cliente_J3")

                fila_actual = 2
                for cliente_j3 in bloque["clientes_j3"]:
                    ws.cell(
                        row=fila_actual,
                        column=columna_actual,
                        value=bloque["cliente_j1"],
                    )
                    ws.cell(
                        row=fila_actual,
                        column=columna_actual + 1,
                        value=bloque["nombre"],
                    )
                    ws.cell(
                        row=fila_actual,
                        column=columna_actual + 2,
                        value=bloque.get("nombre2", ""),
                    )
                    ws.cell(
                        row=fila_actual, column=columna_actual + 3, value=cliente_j3
                    )
                    fila_actual += 1

                columna_actual += 6

            print(f"[DSD] Guardando archivo en: {archivo_salida}")
            wb.save(archivo_salida)
            archivo_generado_exitosamente = True
            print(f"[DSD] Archivo guardado exitosamente con openpyxl")

        except ImportError as e:
            print(f"[DSD] openpyxl no disponible ({e}), intentando con xlsxwriter...")

            try:
                import xlsxwriter

                print(f"[DSD] Generando archivo con xlsxwriter...")
                workbook = xlsxwriter.Workbook(archivo_salida)
                worksheet = workbook.add_worksheet("Solicitantes")

                columna_actual = 0

                for idx_bloque, bloque in enumerate(bloques):

                    worksheet.write(0, columna_actual, "Cliente_J1")
                    worksheet.write(0, columna_actual + 1, "Nombre_Solicitante")
                    worksheet.write(0, columna_actual + 2, "Nombre_2")
                    worksheet.write(0, columna_actual + 3, "Cliente_J3")

                    fila_actual = 1
                    for cliente_j3 in bloque["clientes_j3"]:
                        worksheet.write(
                            fila_actual, columna_actual, bloque["cliente_j1"]
                        )
                        worksheet.write(
                            fila_actual, columna_actual + 1, bloque["nombre"]
                        )
                        worksheet.write(
                            fila_actual, columna_actual + 2, bloque.get("nombre2", "")
                        )
                        worksheet.write(fila_actual, columna_actual + 3, cliente_j3)
                        fila_actual += 1

                    columna_actual += 6

                workbook.close()
                archivo_generado_exitosamente = True
                print(f"[DSD] Archivo guardado exitosamente con xlsxwriter")
            except ImportError as e2:
                print(
                    f"[DSD] xlsxwriter no disponible ({e2}), intentando con pandas..."
                )

                try:
                    max_filas = (
                        max(len(bloque["clientes_j3"]) for bloque in bloques) + 1
                    )

                    datos_por_columna = {}
                    columna_actual = 0

                    for idx_bloque, bloque in enumerate(bloques):
                        datos_por_columna[columna_actual] = ["Cliente_J1"] + [
                            bloque["cliente_j1"]
                        ] * len(bloque["clientes_j3"])
                        datos_por_columna[columna_actual + 1] = [
                            "Nombre_Solicitante"
                        ] + [bloque["nombre"]] * len(bloque["clientes_j3"])
                        datos_por_columna[columna_actual + 2] = ["Nombre_2"] + [
                            bloque.get("nombre2", "")
                        ] * len(bloque["clientes_j3"])
                        datos_por_columna[columna_actual + 3] = ["Cliente_J3"] + bloque[
                            "clientes_j3"
                        ]

                        datos_por_columna[columna_actual + 4] = [""] * max_filas
                        datos_por_columna[columna_actual + 5] = [""] * max_filas

                        columna_actual += 6

                    for col_idx in datos_por_columna:
                        while len(datos_por_columna[col_idx]) < max_filas:
                            datos_por_columna[col_idx].append("")

                    df_resultado = pd.DataFrame(datos_por_columna)

                    print(f"[DSD] Guardando archivo con pandas...")
                    with pd.ExcelWriter(archivo_salida, engine=ENGINE_XLS) as writer:
                        df_resultado.to_excel(
                            writer, sheet_name="Solicitantes", index=False, header=False
                        )
                    archivo_generado_exitosamente = True
                    print(f"[DSD] Archivo guardado exitosamente con pandas")
                except Exception as e3:
                    print(f"[DSD] Error con pandas: {e3}")
                    raise Exception(
                        f"No se pudo generar el archivo Excel. Errores: openpyxl={e}, xlsxwriter={e2}, pandas={e3}"
                    )

        except Exception as e_gen:
            print(f"[DSD] Error general generando archivo: {e_gen}")
            import traceback

            traceback.print_exc()
            raise Exception(f"Error al generar el archivo Excel: {str(e_gen)}")

        if not archivo_generado_exitosamente:
            raise Exception("El archivo no se generó correctamente")

        if not os.path.exists(archivo_salida):
            raise Exception(
                f"El archivo se generó pero no se encuentra en: {archivo_salida}"
            )

        print(
            f"[DSD] Archivo verificado: {archivo_salida} (tamaño: {os.path.getsize(archivo_salida)} bytes)"
        )

        total_filas = sum(len(bloque["clientes_j3"]) for bloque in bloques)

        progreso_dsd["estado"] = "completado"
        progreso_dsd["etapa"] = "Proceso completado exitosamente"
        progreso_dsd["progreso"] = 100
        progreso_dsd["total_filas"] = total_filas
        progreso_dsd["archivo_generado"] = archivo_salida

        if archivo_dsd_subido and os.path.exists(archivo_dsd_subido):
            try:
                os.remove(archivo_dsd_subido)
                archivo_dsd_subido = None
                print(f"[DSD] Archivo temporal eliminado")
            except Exception as e:
                print(f"[DSD] No se pudo eliminar archivo temporal: {e}")

    except Exception as e:
        progreso_dsd["estado"] = "error"
        progreso_dsd["etapa"] = "Error"
        progreso_dsd["mensaje"] = str(e)
        print(f"[DSD] Error: {e}")
        import traceback

        traceback.print_exc()

        if archivo_dsd_subido and os.path.exists(archivo_dsd_subido):
            try:
                os.remove(archivo_dsd_subido)
                archivo_dsd_subido = None
            except Exception:
                pass


@app.route("/api/dsd/upload", methods=["POST"])
def api_dsd_upload():
    """Endpoint para subir el archivo Base Jerarquia.xlsx"""
    global archivo_dsd_subido

    if "file" not in request.files:
        return jsonify({"error": "No se envió ningún archivo"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No se seleccionó ningún archivo"}), 400

    if not (file.filename.lower().endswith((".xlsx", ".xls", ".xlsm"))):
        return (
            jsonify({"error": "El archivo debe ser un Excel (.xlsx, .xls, .xlsm)"}),
            400,
        )

    try:
        for old_file in os.listdir(DSD_TEMP_PATH):
            try:
                os.remove(os.path.join(DSD_TEMP_PATH, old_file))
            except Exception:
                pass

        filename = f"Base_Jerarquia_{uuid.uuid4().hex[:8]}.xlsx"
        filepath = os.path.join(DSD_TEMP_PATH, filename)
        file.save(filepath)

        archivo_dsd_subido = filepath
        print(f"[DSD] Archivo subido: {filename}")

        return jsonify({"ok": True, "filename": file.filename})
    except Exception as e:
        print(f"[DSD] Error al subir archivo: {e}")
        return jsonify({"error": f"Error al guardar el archivo: {str(e)}"}), 500


@app.route("/api/dsd/iniciar", methods=["POST"])
def api_dsd_iniciar():
    global archivo_dsd_subido

    if progreso_dsd.get("estado") == "procesando":
        return jsonify({"error": "Ya hay un proceso en ejecución"}), 400

    if not archivo_dsd_subido or not os.path.exists(archivo_dsd_subido):
        return (
            jsonify(
                {
                    "error": "No hay archivo subido. Por favor, sube el archivo Base Jerarquia.xlsx primero."
                }
            ),
            400,
        )

    thread = threading.Thread(target=_run_dsd, daemon=True)
    thread.start()
    return jsonify({"ok": True})


@app.route("/api/dsd/estado", methods=["GET"])
def api_dsd_estado():
    return jsonify(progreso_dsd)


@app.route("/api/dsd/descargar", methods=["GET"])
def api_dsd_descargar():
    archivo = progreso_dsd.get("archivo_generado")
    if not archivo or not os.path.exists(archivo):
        return jsonify({"error": "Archivo no disponible"}), 404
    carpeta = os.path.dirname(archivo)
    nombre = os.path.basename(archivo)
    return send_from_directory(carpeta, nombre, as_attachment=True)


@app.route("/facturas")
def facturas_index():
    """
    Carga la interfaz de facturas (HTML en Facturas/cod_facturas/index.html)
    """
    return send_from_directory(FACTURAS_CODIGO_PATH, "index.html")


@app.route("/facturas/styles.css")
def facturas_css():
    return send_from_directory(
        FACTURAS_CODIGO_PATH, "styles.css", mimetype="text/css", max_age=0
    )


@app.route("/facturas/<path:filename>")
def facturas_static(filename: str):
    return send_from_directory(FACTURAS_CODIGO_PATH, filename)


@app.route("/upload", methods=["POST", "OPTIONS"])
def upload_files():
    """
    Endpoint de subida de facturas usado por la interfaz de FactuVal.
    NO actualiza Excel de portafolios, solo procesa PDFs y genera Excel de facturas.
    """
    if request.method == "OPTIONS":
        return "", 200

    if main_facturas is None or procesar_facturas is None:
        return (
            jsonify({"status": "error", "message": "Módulo de facturas no disponible"}),
            500,
        )

    modo = request.form.get("modo", "acumular")
    print(f"[UPLOAD] Modo de procesamiento: {modo}")

    if modo == "separado" or request.form.get("limpiar") == "1":
        for old_file in os.listdir(FACTURAS_PATH):
            old_path = os.path.join(FACTURAS_PATH, old_file)
            try:
                os.remove(old_path)
            except Exception as e:
                print(f"[UPLOAD] No se pudo eliminar {old_path}: {e}")
    else:
        print(
            f"[UPLOAD] Modo acumular: se mantienen archivos existentes en {FACTURAS_PATH}"
        )

    files = request.files.getlist("files")
    if not files:
        return jsonify({"status": "error", "message": "No se enviaron archivos"}), 400

    saved: list[str] = []
    archivos_rechazados_locales: list[dict] = []

    for f in files:
        filename = f.filename
        dest = os.path.join(FACTURAS_PATH, filename)
        f.save(dest)

        es_valida, mensaje = main_facturas._es_factura_valida(dest)
        if not es_valida:
            archivos_rechazados_locales.append({"archivo": filename, "razon": mensaje})
            destino_rechazado = os.path.join(FACTURAS_RECHAZADOS, filename)
            try:
                main_facturas.mover_archivo_seguro(dest, destino_rechazado)
            except Exception as e:
                print(f"[UPLOAD] Error al mover archivo rechazado: {e}")
        else:
            saved.append(filename)
            print(f"[UPLOAD] Archivo válido: {filename}")

    if not saved:
        return jsonify(
            {
                "status": "ok",
                "saved": saved,
                "processing_started": False,
                "run_id": None,
                "archivos_rechazados": archivos_rechazados_locales,
                "message": "No hay archivos válidos para procesar",
            }
        )

    run_id = uuid.uuid4().hex
    global resultados_globales
    resultados_globales = {
        "status": "processing",
        "tiempo_inicio": time.time(),
        "tiempo_actual": 0,
        "run_id": run_id,
    }

    thread = threading.Thread(
        target=run_procesamiento_facturas, args=(modo,), daemon=True
    )
    thread.start()

    return jsonify(
        {
            "status": "ok",
            "saved": saved,
            "processing_started": True,
            "run_id": run_id,
            "archivos_rechazados": archivos_rechazados_locales,
        }
    )


def run_procesamiento_facturas(modo: str):
    """
    Hilo en background que llama a procesar_facturas(modo) del módulo main_facturas.
    """
    print("[SERVER] Iniciando procesamiento de facturas...")

    if procesar_facturas is None:
        print("[SERVER] ERROR: procesar_facturas no está disponible")
        return

    resultado = procesar_facturas(modo)

    global resultados_globales

    if resultado is None:
        print("[SERVER] ERROR: procesar_facturas devolvió None")
        resultados_globales = {
            "status": "error",
            "tiempos": [],
            "total": "0 seg",
            "archivos_rechazados": [],
            "facturas_con_errores": [],
            "modo": modo,
            "excel": "error.xlsx",
            "run_id": resultados_globales.get("run_id"),
            "facturas_procesadas": 0,
            "facturas_nuevas": 0,
            "facturas_duplicadas": 0,
            "error": "Error interno: No se pudo procesar las facturas",
        }
        return

    resultados_globales = {
        "status": "done",
        "tiempos": resultado.get("tiempos", []),
        "total": formatear_tiempo(resultado.get("total_seconds", 0)),
        "archivos_rechazados": resultado.get("archivos_rechazados", []),
        "facturas_con_errores": resultado.get("facturas_con_errores", []),
        "modo": modo,
        "excel": os.path.basename(resultado.get("excel_path") or "procesadas.xlsx"),
        "run_id": resultados_globales.get("run_id"),
        "facturas_procesadas": resultado.get("facturas_procesadas", 0),
        "facturas_nuevas": resultado.get(
            "facturas_nuevas", resultado.get("facturas_procesadas", 0)
        ),
        "facturas_duplicadas": resultado.get("facturas_duplicadas", 0),
    }

    print(
        f"[SERVER] Procesamiento de facturas terminado en {resultados_globales['total']}"
    )


@app.route("/resultado", methods=["GET"])
def resultado():
    if resultados_globales.get("status") == "done":
        excel_filename = resultados_globales.get("excel", "procesadas.xlsx")
        excel_path = os.path.join(FACTURAS_ROOT, excel_filename)
        if os.path.exists(excel_path):
            print(f"[SERVER] Devolviendo resultados y {excel_path}")
        return jsonify(
            {
                "status": "done",
                "tiempos": resultados_globales.get("tiempos", []),
                "total": resultados_globales.get("total", ""),
                "excel": excel_filename,
                "archivos_rechazados": resultados_globales.get(
                    "archivos_rechazados", []
                ),
                "facturas_con_errores": resultados_globales.get(
                    "facturas_con_errores", []
                ),
                "modo": resultados_globales.get("modo", "acumular"),
                "run_id": resultados_globales.get("run_id"),
                "facturas_procesadas": resultados_globales.get(
                    "facturas_procesadas", 0
                ),
                "facturas_nuevas": resultados_globales.get("facturas_nuevas", 0),
                "facturas_duplicadas": resultados_globales.get(
                    "facturas_duplicadas", 0
                ),
            }
        )

    tiempo_inicio = resultados_globales.get("tiempo_inicio", time.time())
    tiempo_actual = time.time() - tiempo_inicio
    tiempo_formateado = formatear_tiempo(tiempo_actual)
    print(
        f"[SERVER] Aún procesando facturas... Tiempo transcurrido: {tiempo_formateado}"
    )
    return jsonify(
        {
            "status": "processing",
            "tiempo_actual": tiempo_formateado,
            "tiempo_segundos": tiempo_actual,
            "run_id": resultados_globales.get("run_id"),
        }
    )


@app.route("/descargar_excel", methods=["GET"])
def descargar_excel():
    filename = request.args.get("file", "procesadas.xlsx")
    excel_path = os.path.join(FACTURAS_ROOT, filename)
    if os.path.exists(excel_path):
        print(f"[SERVER] Enviando archivo: {filename}")
        return send_from_directory(FACTURAS_ROOT, filename, as_attachment=True)
    return jsonify({"status": "error", "message": "El archivo aún no está disponible"})


@app.route("/<path:filename>")
def serve_static(filename: str):
    """
    Cualquier otro archivo estático se sirve desde la carpeta Web.
    """
    return send_from_directory(CODIGOS_PRINCIPALES_PATH, filename)


if __name__ == "__main__":
    print("=" * 60)
    print("SERVIDOR UNIFICADO - POSTOBON S.A. (Automatizaciones_postobon)")
    print("=" * 60)
    print("Servidor ejecutándose en:")
    print("  http://localhost:5000")
    print("  http://127.0.0.1:5000")
    print("=" * 60)
    print("Aplicaciones disponibles:")
    print("  - Menú Principal:           http://localhost:5000/")
    print("  - FactuVal:                 http://localhost:5000/facturas")
    print("  - Consulta de pedidos DSD:  http://localhost:5000/dsd")
    print("=" * 60)

    app.run(debug=False, host="0.0.0.0", port=5000, threaded=True)
