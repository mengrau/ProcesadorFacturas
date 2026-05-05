from __future__ import annotations

import logging
import os
import shutil
import time
import zipfile
from collections.abc import Callable
from pathlib import Path
from typing import Any

import openpyxl

from facturas_app.services.invoice_parser import format_number, normalize_number

logger = logging.getLogger(__name__)

InvoiceRow = dict[str, Any]

EXCEL_HEADERS = [
    "ID",
    "Número Factura",
    "NIT Cliente",
    "Cod Cliente",
    "Nombre Cliente",
    "Fecha Generación",
    "Fecha Expedición",
    "Referencia",
    "Producto",
    "UMV",
    "Unidades",
    "Precio Base U",
    "IVA",
    "Total",
    "Estado",
]


class InvoiceExcelRepository:
    """Repository responsible for reading and writing the invoice Excel file."""

    def __init__(
        self,
        excel_path: str | Path,
        *,
        message_callback: Callable[[str], None] | None = None,
    ) -> None:
        self.excel_path = Path(excel_path)
        self._message_callback = message_callback

    def _message(self, text: str) -> None:
        if self._message_callback is not None:
            self._message_callback(text)

    @staticmethod
    def _row_value(row: Any, index: int, default: str = "") -> str:
        try:
            value = row[index]
        except Exception:
            return default
        return str(value) if value else default

    @staticmethod
    def _pandas_value(row: Any, index: int, default: str = "") -> str:
        try:
            import pandas as pd

            value = row.iloc[index]
            if pd.notna(value):
                return str(value)
        except Exception:
            return default
        return default

    @classmethod
    def _record_from_openpyxl_row(cls, row: tuple[Any, ...]) -> InvoiceRow:
        return {
            "id": cls._row_value(row, 0),
            "numero_factura": cls._row_value(row, 1),
            "nit_cliente": cls._row_value(row, 2),
            "cod_cliente": cls._row_value(row, 3),
            "nombre_cliente": cls._row_value(row, 4),
            "fecha_generacion": cls._row_value(row, 5),
            "fecha_expedicion": cls._row_value(row, 6),
            "referencia": cls._row_value(row, 7),
            "productos": cls._row_value(row, 8),
            "umv": cls._row_value(row, 9),
            "unidades": cls._row_value(row, 10),
            "precio_base_unitario": cls._row_value(row, 11),
            "iva": cls._row_value(row, 12, "0.00"),
            "total": cls._row_value(row, 13),
            "estado": cls._row_value(row, 14, "OK"),
        }

    @classmethod
    def _record_from_pandas_row(cls, row: Any) -> InvoiceRow:
        return {
            "id": cls._pandas_value(row, 0),
            "numero_factura": cls._pandas_value(row, 1),
            "nit_cliente": cls._pandas_value(row, 2),
            "cod_cliente": cls._pandas_value(row, 3),
            "nombre_cliente": cls._pandas_value(row, 4),
            "fecha_generacion": cls._pandas_value(row, 5),
            "fecha_expedicion": cls._pandas_value(row, 6),
            "referencia": cls._pandas_value(row, 7),
            "productos": cls._pandas_value(row, 8),
            "umv": cls._pandas_value(row, 9),
            "unidades": cls._pandas_value(row, 10),
            "precio_base_unitario": cls._pandas_value(row, 11),
            "iva": cls._pandas_value(row, 12, "0.00"),
            "total": cls._pandas_value(row, 13),
            "estado": cls._pandas_value(row, 14, "OK"),
        }

    def load_existing_invoice_numbers(self) -> set[str]:
        """Load already processed invoice numbers from the Excel file."""
        existing_invoices: set[str] = set()
        excel_path = str(self.excel_path)

        if not os.path.exists(excel_path):
            return existing_invoices

        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 1 and row[1]:
                    invoice_number = str(row[1]).strip()
                    if invoice_number and invoice_number != "None":
                        existing_invoices.add(invoice_number)

            self._message(
                f"    [INFO] Cargadas {len(existing_invoices)} facturas existentes del Excel"
            )
            return existing_invoices

        except Exception as exc:
            self._message(
                f"    [WARN] Error al cargar facturas existentes con openpyxl: {exc}"
            )

        try:
            import pandas as pd

            df = pd.read_excel(excel_path, engine="openpyxl")
            if not df.empty and len(df.columns) >= 2:
                for _, row in df.iterrows():
                    if pd.notna(row.iloc[1]):
                        invoice_number = str(row.iloc[1]).strip()
                        if invoice_number and invoice_number != "None":
                            existing_invoices.add(invoice_number)
                self._message(
                    f"    [INFO] Cargadas {len(existing_invoices)} facturas existentes usando pandas"
                )
                return existing_invoices
        except Exception as exc:
            self._message(
                f"    [WARN] Error al cargar facturas existentes con pandas: {exc}"
            )

        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 1 and row[1]:
                    invoice_number = str(row[1]).strip()
                    if invoice_number and invoice_number != "None":
                        existing_invoices.add(invoice_number)

            wb.close()
            self._message(
                f"    [INFO] Cargadas {len(existing_invoices)} facturas existentes en modo read_only"
            )
            return existing_invoices

        except Exception as exc:
            self._message(
                f"    [WARN] Error al cargar facturas existentes en modo read_only: {exc}"
            )

        self._message(
            "    [WARN] No se pudieron cargar facturas existentes, se procesarán todas como nuevas"
        )
        return existing_invoices

    def recover_existing_data(self) -> list[InvoiceRow]:
        """Try to recover rows from an existing Excel file using multiple methods."""
        recovered_data: list[InvoiceRow] = []
        excel_path = str(self.excel_path)

        try:
            import pandas as pd

            df = pd.read_excel(excel_path, engine="openpyxl")
            if not df.empty and len(df.columns) >= 14:
                self._message(
                    f"    [INFO] Recuperando datos con pandas: {len(df)} filas encontradas"
                )
                for _, row in df.iterrows():
                    if pd.notna(row.iloc[1]):
                        recovered_data.append(self._record_from_pandas_row(row))
                self._message(
                    f"    [OK] Se recuperaron {len(recovered_data)} registros del Excel existente"
                )
                return recovered_data
        except Exception as exc:
            self._message(f"    [WARN] Método pandas falló: {exc}")

        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            ws = wb.active

            for cell in ws[1]:
                str(cell.value) if cell.value else ""

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 1 and row[1]:
                    recovered_data.append(self._record_from_openpyxl_row(row))
            wb.close()
            if recovered_data:
                self._message(
                    f"    [OK] Se recuperaron {len(recovered_data)} registros con openpyxl"
                )
                return recovered_data
        except Exception as exc:
            self._message(f"    [WARN] Método openpyxl falló: {exc}")

        return recovered_data

    @staticmethod
    def _deduplication_key(record: InvoiceRow) -> tuple[str, str, str, str]:
        return (
            str(record.get("numero_factura", "")),
            str(record.get("referencia", "")),
            str(record.get("productos", "")),
            normalize_number(str(record.get("total", ""))),
        )

    @staticmethod
    def _append_headers(ws: Any) -> None:
        ws.append(EXCEL_HEADERS)

    def _load_existing_rows_for_accumulated_mode(
        self,
    ) -> list[InvoiceRow]:
        excel_path = str(self.excel_path)
        existing_data: list[InvoiceRow] = []
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

            self._message(
                "    [INFO] Modo acumular: Recuperando datos existentes del Excel"
            )
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 1 and row[1]:
                    existing_data.append(self._record_from_openpyxl_row(row))

            self._message(
                f"    [INFO] Modo acumular: {len(existing_data)} registros existentes recuperados"
            )
            wb.close()
            return existing_data

        except (
            EOFError,
            openpyxl.utils.exceptions.InvalidFileException,
            zipfile.BadZipFile,
            Exception,
        ) as exc:
            self._message(
                f"    [WARN] Error al cargar Excel existente con openpyxl: {exc}"
            )
            self._message(
                "    [INFO] Intentando recuperar datos antes de recrear el archivo..."
            )

            existing_data = self.recover_existing_data()

            if existing_data:
                self._message(
                    f"    [OK] Se recuperaron {len(existing_data)} registros antes de recrear el archivo"
                )
                try:
                    backup_path = excel_path + f".backup_{int(time.time())}"
                    shutil.copy2(excel_path, backup_path)
                    self._message(
                        f"    [INFO] Backup creado: {os.path.basename(backup_path)}"
                    )
                except Exception as backup_error:
                    self._message(f"    [WARN] No se pudo crear backup: {backup_error}")

                try:
                    os.remove(excel_path)
                    self._message("    [OK] Archivo corrupto eliminado")
                except Exception as remove_error:
                    self._message(
                        f"    [WARN] No se pudo eliminar archivo corrupto: {remove_error}"
                    )
            else:
                self._message(
                    "    [WARN] No se pudieron recuperar datos del archivo corrupto"
                )
                try:
                    backup_path = excel_path + f".backup_{int(time.time())}"
                    shutil.copy2(excel_path, backup_path)
                    self._message(
                        f"    [INFO] Backup creado antes de eliminar: {os.path.basename(backup_path)}"
                    )
                except Exception:
                    pass
                try:
                    os.remove(excel_path)
                except Exception:
                    pass

            self._message("    [INFO] Creando nuevo archivo Excel...")
            return existing_data

    def _create_new_workbook(self, mode: str) -> tuple[Any, Any]:
        wb = openpyxl.Workbook()
        ws = wb.active
        self._append_headers(ws)
        if mode == "separado":
            self._message("    [INFO] Modo separado: Creando nuevo archivo Excel")
        else:
            self._message(
                "    [INFO] Modo acumular: Creando nuevo archivo Excel (no existe archivo previo)"
            )
        return wb, ws

    def save(self, records: list[InvoiceRow], mode: str = "acumular") -> str:
        """Save invoice rows into the configured Excel file preserving legacy format."""
        excel_path = str(self.excel_path)
        existing_data: list[InvoiceRow] = []
        save_started_at = time.perf_counter()
        logger.info(
            "Excel write start. mode=%s path=%s incoming_records=%s",
            mode,
            excel_path,
            len(records),
        )

        self.excel_path.parent.mkdir(parents=True, exist_ok=True)

        if mode == "acumular" and os.path.exists(excel_path):
            existing_data = self._load_existing_rows_for_accumulated_mode()
            self._message(
                f"    [INFO] Agregando {len(records)} nuevos registros al Excel existente"
            )
            wb, ws = self._create_new_workbook(mode)
        else:
            wb, ws = self._create_new_workbook(mode)

        all_data: dict[tuple[str, str, str, str], InvoiceRow] = {}

        for record in existing_data:
            key = self._deduplication_key(record)
            if key not in all_data:
                all_data[key] = record

        for record in records:
            key = self._deduplication_key(record)
            all_data[key] = record

        for record in all_data.values():
            base_price = record.get("precio_base_unitario", "")
            unit = str(record.get("umv", ""))

            if unit.upper() == "BOL":
                base_price = str(base_price).replace(",", ".")
            else:
                base_price = format_number(str(base_price))

            ws.append(
                [
                    record.get("id", ""),
                    record.get("numero_factura", ""),
                    record.get("nit_cliente", ""),
                    record.get("cod_cliente", ""),
                    record.get("nombre_cliente", ""),
                    record.get("fecha_generacion", ""),
                    record.get("fecha_expedicion", ""),
                    record.get("referencia", ""),
                    record.get("productos", ""),
                    record.get("umv", ""),
                    record.get("unidades", ""),
                    base_price,
                    record.get("iva", "0.00"),
                    format_number(str(record.get("total", ""))),
                    record.get("estado", ""),
                ]
            )

        logger.info(
            "Excel write rows staged. mode=%s staged_rows=%s", mode, len(all_data)
        )
        excel_save_started_at = time.perf_counter()
        logger.info("Excel write save start. path=%s", excel_path)

        wb.save(excel_path)
        logger.info(
            "Excel write save done. path=%s elapsed=%.3fs",
            excel_path,
            time.perf_counter() - excel_save_started_at,
        )
        logger.info(
            "Excel write done. mode=%s total_elapsed=%.3fs total_rows=%s",
            mode,
            time.perf_counter() - save_started_at,
            len(all_data),
        )
        self._message(
            f"    [OK] Excel guardado con {len(all_data)} registros totales ({len(existing_data)} existentes + {len(records)} nuevos)"
        )
        return excel_path
