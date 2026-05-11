from __future__ import annotations

import unicodedata
from collections.abc import Callable
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook

from facturas_app.config import Settings, get_settings


ProgressCallback = Callable[[dict[str, Any]], None]


def normalize_text(value: str) -> str:
    """Normalize column names for resilient matching."""
    if not isinstance(value, str):
        return ""
    value = value.strip().lower()
    return "".join(
        char
        for char in unicodedata.normalize("NFD", value)
        if unicodedata.category(char) != "Mn"
    )


class DsdService:
    """Build the DSD solicitantes workbook from Base Jerarquia data."""

    def __init__(self, settings: Settings | None = None) -> None:
        self.settings = settings or get_settings()

    @staticmethod
    def _update(callback: ProgressCallback | None, **data: Any) -> None:
        if callback is not None:
            callback(data)

    @staticmethod
    def _find_columns(df: pd.DataFrame) -> tuple[str, str, str | None, str]:
        columns = {normalize_text(column): column for column in df.columns}

        col_j1 = None
        col_nombre = None
        col_nombre2 = None
        col_j3 = None

        for normalized, original in columns.items():
            if "cliente" in normalized and "j1" in normalized:
                col_j1 = original
                break

        for normalized, original in columns.items():
            if normalized == "nombre" or (
                normalized.startswith("nombre")
                and "2" not in normalized
                and "nombre2" not in normalized
            ):
                col_nombre = original
                break

        for normalized, original in columns.items():
            if "nombre" in normalized and (
                "2" in normalized
                or normalized == "nombre2"
                or normalized.startswith("nombre2")
            ):
                col_nombre2 = original
                break

        for normalized, original in columns.items():
            if "cliente" in normalized and "j3" in normalized:
                col_j3 = original
                break

        if not col_j1:
            raise ValueError("No se encontro la columna 'Cliente J1' en el archivo")
        if not col_nombre:
            raise ValueError("No se encontro la columna 'Nombre' en el archivo")
        if not col_j3:
            raise ValueError("No se encontro la columna 'Cliente J3' en el archivo")

        return col_j1, col_nombre, col_nombre2, col_j3

    def _load_dataframe(self, source_path: Path) -> tuple[pd.DataFrame, str]:
        excel_file = pd.ExcelFile(source_path)

        for sheet_name in excel_file.sheet_names:
            try:
                df = pd.read_excel(source_path, sheet_name=sheet_name)
                columns = {normalize_text(column): column for column in df.columns}
                has_j1 = any("cliente" in item and "j1" in item for item in columns)
                has_name = "nombre" in columns
                has_j3 = any("cliente" in item and "j3" in item for item in columns)
                if has_j1 and has_name and has_j3:
                    return df, sheet_name
            except Exception:
                continue

        fallback_sheet = (
            excel_file.sheet_names[0] if excel_file.sheet_names else "Sheet1"
        )
        return pd.read_excel(source_path, sheet_name=0), fallback_sheet

    @staticmethod
    def _build_blocks(
        df: pd.DataFrame,
        *,
        col_j1: str,
        col_nombre: str,
        col_nombre2: str | None,
        col_j3: str,
        callback: ProgressCallback | None,
    ) -> list[dict[str, Any]]:
        clientes_j1 = df[col_j1].dropna().unique()
        clientes_j1 = [client for client in clientes_j1 if str(client).strip()]

        if not clientes_j1:
            raise ValueError("No se encontraron valores en la columna 'Cliente J1'")

        blocks: list[dict[str, Any]] = []
        total_clients = len(clientes_j1)

        for index, cliente_j1 in enumerate(clientes_j1):
            DsdService._update(
                callback,
                etapa=(
                    f"Procesando cliente J1: {cliente_j1} "
                    f"({index + 1}/{total_clients})"
                ),
                progreso=40 + int((index / total_clients) * 50),
                clientes_procesados=index + 1,
            )

            filtered = df[df[col_j1] == cliente_j1].copy()
            if filtered.empty:
                continue

            nombre = filtered[col_nombre].iloc[0] if len(filtered) > 0 else ""
            nombre2 = ""
            if col_nombre2:
                nombre2 = filtered[col_nombre2].iloc[0] if len(filtered) > 0 else ""

            clientes_j3 = filtered[col_j3].dropna()
            unique_j3: list[str] = []
            seen: set[str] = set()
            for cliente_j3 in clientes_j3:
                value = str(cliente_j3).strip()
                if value and value not in seen:
                    unique_j3.append(value)
                    seen.add(value)

            if unique_j3:
                blocks.append(
                    {
                        "cliente_j1": str(cliente_j1).strip(),
                        "nombre": str(nombre).strip() if nombre else "",
                        "nombre2": str(nombre2).strip() if nombre2 else "",
                        "clientes_j3": unique_j3,
                    }
                )

        if not blocks:
            raise ValueError(
                "No se generaron datos. "
                "Verifica que el archivo tenga informacion valida."
            )

        return blocks

    def _write_output(self, blocks: list[dict[str, Any]]) -> Path:
        self.settings.ruta_salida_dsd.mkdir(parents=True, exist_ok=True)
        output_path = self.settings.ruta_salida_dsd / "Solicitantes_SAP.xlsx"

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Solicitantes"

        current_column = 1
        for block in blocks:
            worksheet.cell(row=1, column=current_column, value="Cliente_J1")
            worksheet.cell(row=1, column=current_column + 1, value="Nombre_Solicitante")
            worksheet.cell(row=1, column=current_column + 2, value="Nombre_2")
            worksheet.cell(row=1, column=current_column + 3, value="Cliente_J3")

            current_row = 2
            for cliente_j3 in block["clientes_j3"]:
                worksheet.cell(
                    row=current_row,
                    column=current_column,
                    value=block["cliente_j1"],
                )
                worksheet.cell(
                    row=current_row,
                    column=current_column + 1,
                    value=block["nombre"],
                )
                worksheet.cell(
                    row=current_row,
                    column=current_column + 2,
                    value=block["nombre2"],
                )
                worksheet.cell(
                    row=current_row,
                    column=current_column + 3,
                    value=cliente_j3,
                )
                current_row += 1

            current_column += 6

        workbook.save(output_path)
        return output_path

    def process(
        self,
        source_path: Path,
        *,
        progress_callback: ProgressCallback | None = None,
    ) -> dict[str, Any]:
        """Process a Base Jerarquia workbook and return the generated file info."""
        if not source_path.exists():
            raise FileNotFoundError(f"No se encontro el archivo: {source_path}")

        self._update(
            progress_callback,
            estado="procesando",
            etapa=f"Abriendo archivo: {source_path.name}...",
            progreso=10,
        )

        df, sheet_name = self._load_dataframe(source_path)

        self._update(
            progress_callback,
            etapa=f"Archivo abierto (hoja: {sheet_name}). Buscando columnas...",
            progreso=20,
        )

        col_j1, col_nombre, col_nombre2, col_j3 = self._find_columns(df)

        self._update(
            progress_callback,
            etapa="Columnas encontradas. Obteniendo clientes unicos...",
            progreso=30,
        )

        blocks = self._build_blocks(
            df,
            col_j1=col_j1,
            col_nombre=col_nombre,
            col_nombre2=col_nombre2,
            col_j3=col_j3,
            callback=progress_callback,
        )

        self._update(
            progress_callback,
            etapa="Generando archivo de salida...",
            progreso=90,
        )

        output_path = self._write_output(blocks)
        total_rows = sum(len(block["clientes_j3"]) for block in blocks)

        return {
            "archivo_generado": str(output_path),
            "total_filas": total_rows,
        }
